"""
Sevak Joda - Member Registration App
Saves member data to Google Sheets, captures photo via webcam/scanner.
"""

import sys
import os
import re
import shutil
import textwrap
from datetime import datetime

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas as pdfcanvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.utils import ImageReader
from marathi_text_render import render_text_line

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QFormLayout, QLineEdit, QTextEdit, QPushButton, QLabel, QMessageBox,
    QFileDialog, QStackedWidget, QFrame, QComboBox, QSpinBox, QListWidget,
    QListWidgetItem, QDialog, QScrollArea, QCheckBox, QTableWidget,
    QTableWidgetItem
)
from PyQt5.QtGui import QFont, QPixmap, QDesktopServices
from PyQt5.QtCore import Qt, QUrl, QThread, pyqtSignal

import gspread
from google.oauth2.service_account import Credentials
from scanner_capture import scan_photo
from translator import translate_to_marathi
from marathi_keyboard import MarathiKeyboard
import drive_helper
from network_utils import with_retry
import openpyxl
from legacy_font_converter import convert_legacy_marathi

# ---------- CONFIG ----------
SHEET_ID = "1UfznX39WYRFcl40K32C8zdi7Xhi4WYVn5rpd3emkIUk"
CREDENTIALS_FILE = "credentials.json"
PHOTOS_DIR = "photos"
MARATHI_FONT = "Noto Sans Devanagari"   # install this font on the OS for best rendering
ENGLISH_FONT = "Arial"
SIDEBAR_WIDTH_EXPANDED = 180
SIDEBAR_WIDTH_COLLAPSED = 48
SHEET_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit"
DRIVE_FOLDER_URL = f"https://drive.google.com/drive/folders/{drive_helper.DRIVE_FOLDER_ID}"
WARI_SHEET_NAME = "Wari Attendees"
WARI_OPTIONS = [
    ("Chaitra", "चैत्र"),
    ("Aashadh", "आषाढ"),
    ("Karthik", "कार्तिक"),
    ("Magh", "माघ"),
    ("Adhik", "अधिक"),
]
# -----------------------------

os.makedirs(PHOTOS_DIR, exist_ok=True)

# Devanagari font for PDF generation (Helvetica can't render Marathi glyphs).
# sys._MEIPASS is where PyInstaller extracts bundled data files at runtime
# for a frozen exe; fall back to the script's own directory when running
# from source.
_APP_DIR = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
PDF_FONT_NAME = "NotoSansDevanagari"
_PDF_FONT_PATH = os.path.join(_APP_DIR, "fonts", "NotoSansDevanagari-Regular.ttf")
try:
    pdfmetrics.registerFont(TTFont(PDF_FONT_NAME, _PDF_FONT_PATH))
except Exception:
    PDF_FONT_NAME = "Helvetica"  # fallback: Marathi text will render as boxes/blank

# Bilingual text for all app chrome (nav, headers, buttons, form labels).
# key -> (english, marathi)
LABELS = {
    "window_title": ("Sevak Joda - Member Registration", "सेवक जोडा - सदस्य नोंदणी"),
    "nav_add_member": ("Add Member", "सेवक जोडा"),
    "nav_edit_member": ("Edit Sevak", "सेवक संपादित करा"),
    "nav_vari_member": ("Wari Attendees", "वारी उपस्थित सेवक"),
    "lang_toggle": ("मराठी", "English"),
    "open_sheet": ("📊 Open Google Sheet", "📊 गूगल शीट उघडा"),
    "open_drive": ("🖼 Open Photos Drive Folder", "🖼 फोटो ड्राइव्ह फोल्डर उघडा"),
    "section_english": ("English Details", "इंग्रजी माहिती"),
    "section_marathi": ("Marathi Details", "मराठी माहिती"),
    "lbl_card_id": ("Card Id:", "कार्ड आयडी:"),
    "lbl_first_en": ("First Name (English):", "पहिले नाव (इंग्रजी):"),
    "lbl_middle_en": ("Middle Name (English):", "मधले नाव (इंग्रजी):"),
    "lbl_last_en": ("Last Name (English):", "आडनाव (इंग्रजी):"),
    "lbl_dob": ("Date of Birth:", "जन्मतारीख:"),
    "lbl_phone": ("Phone Number:", "फोन नंबर:"),
    "lbl_address_en": ("Address (English):", "पत्ता (इंग्रजी):"),
    "lbl_pincode": ("Pincode:", "पिनकोड:"),
    "lbl_first_mr": ("पहिले नाव (Marathi):", "पहिले नाव (मराठी):"),
    "lbl_middle_mr": ("मधले नाव (Marathi):", "मधले नाव (मराठी):"),
    "lbl_last_mr": ("आडनाव (Marathi):", "आडनाव (मराठी):"),
    "lbl_address_mr": ("पत्ता (Marathi):", "पत्ता (मराठी):"),
    "lbl_mukkam_en": ("Mukkam (English):", "मुक्काम (इंग्रजी):"),
    "lbl_post_en": ("Post (English):", "पोस्ट (इंग्रजी):"),
    "lbl_taluka_en": ("Taluka (English):", "तालुका (इंग्रजी):"),
    "lbl_district_en": ("District (English):", "जिल्हा (इंग्रजी):"),
    "lbl_state_en": ("State (English):", "राज्य (इंग्रजी):"),
    "lbl_mukkam_mr": ("मुक्काम (Marathi):", "मुक्काम (मराठी):"),
    "lbl_post_mr": ("पोस्ट (Marathi):", "पोस्ट (मराठी):"),
    "lbl_taluka_mr": ("तालुका (Marathi):", "तालुका (मराठी):"),
    "lbl_jilha_mr": ("जिल्हा (Marathi):", "जिल्हा (मराठी):"),
    "lbl_state_mr": ("राज्य (Marathi):", "राज्य (मराठी):"),
    "btn_translate": ("🌐 Translate English → Marathi", "🌐 इंग्रजी → मराठी भाषांतर"),
    "btn_scan": ("📷 Scan Photo", "📷 फोटो स्कॅन करा"),
    "btn_browse": ("🖼 Select from Local", "🖼 स्थानिक फाइल निवडा"),
    "btn_drive_pick": ("☁ Pick from Drive", "☁ ड्राइव्हमधून निवडा"),
    "drive_picker_title": ("Select Photo from Drive", "ड्राइव्हमधून फोटो निवडा"),
    "drive_no_photos": ("No photos found in the Drive folder.", "ड्राइव्ह फोल्डरमध्ये कोणतेही फोटो सापडले नाहीत."),
    "drive_upload_failed": ("Photo saved locally, but upload to Drive failed", "फोटो स्थानिक पातळीवर जतन झाला, पण ड्राइव्हवर अपलोड अयशस्वी झाले"),
    "btn_save": ("Save Member", "सेवक जतन करा"),
    "no_photo": ("No photo captured", "फोटो नाही"),
    "coming_soon": ("Coming soon", "लवकरच येत आहे"),
    "missing_info_title": ("Missing Info", "माहिती अपूर्ण"),
    "missing_info_body": ("First Name (English) is required.", "पहिले नाव (इंग्रजी) आवश्यक आहे."),
    "success_title": ("Success", "यशस्वी"),
    "success_body": ("Member saved successfully!", "सेवक यशस्वीरित्या जतन झाला!"),
    "error_title": ("Error", "त्रुटी"),
    "lbl_wari_select": ("Select Wari:", "वारी निवडा:"),
    "lbl_wari_year": ("Wari Year:", "वारी वर्ष:"),
    "lbl_search_member": ("Search Member (Name or Card Id):", "सेवक शोधा (नाव किंवा कार्ड आयडी):"),
    "btn_search": ("🔍 Search", "🔍 शोधा"),
    "btn_add_to_wari": ("➕ Add to Wari Attendees", "➕ वारी उपस्थितीत जोडा"),
    "lbl_attendees_list": ("Attendees for selected Wari:", "निवडलेल्या वारीसाठी उपस्थित सेवक:"),
    "no_member_selected_title": ("No Member Selected", "सेवक निवडलेला नाही"),
    "no_member_selected_body": ("Please search and select a member first.", "कृपया आधी सेवक शोधा आणि निवडा."),
    "duplicate_title": ("Already Added", "आधीच जोडले आहे"),
    "duplicate_body": ("This member is already in the attendee list for this Wari/year.",
                        "हा सेवक या वारी/वर्षासाठी आधीच यादीत आहे."),
    "wari_added_body": ("Member added to Wari Attendees!", "सेवक वारी उपस्थितीत जोडला!"),
    "btn_show_list": ("📋 Show List", "📋 यादी दाखवा"),
    "btn_remove_attendee": ("🗑 Remove Selected", "🗑 निवडलेले काढा"),
    "remove_confirm_title": ("Remove Attendee?", "उपस्थित सेवक काढायचा?"),
    "remove_confirm_body": ("Remove this member from the attendee list?",
                            "या सेवकाला उपस्थित यादीतून काढायचे का?"),
    "no_attendee_selected_title": ("No Selection", "निवड नाही"),
    "no_attendee_selected_body": ("Please select an attendee from the list first.",
                                  "कृपया आधी यादीतून एक सेवक निवडा."),
    "remove_success_body": ("Attendee removed.", "सेवक काढला."),
    "status_processing": ("⏳ Processing... please wait", "⏳ प्रक्रिया सुरू आहे... कृपया प्रतीक्षा करा"),
    "status_scanning": ("⏳ Scanning... please wait", "⏳ स्कॅन करत आहे... कृपया प्रतीक्षा करा"),
    "status_saving": ("⏳ Saving... please wait", "⏳ जतन करत आहे... कृपया प्रतीक्षा करा"),
    "status_loading": ("⏳ Loading... please wait", "⏳ लोड करत आहे... कृपया प्रतीक्षा करा"),
    "status_searching": ("⏳ Searching... please wait", "⏳ शोधत आहे... कृपया प्रतीक्षा करा"),
    "status_uploading": ("⏳ Uploading photo... please wait", "⏳ फोटो अपलोड करत आहे... कृपया प्रतीक्षा करा"),
    "status_generating_pdf": ("⏳ Generating PDF... please wait", "⏳ पीडीएफ तयार करत आहे... कृपया प्रतीक्षा करा"),
    "status_removing": ("⏳ Removing... please wait", "⏳ काढत आहे... कृपया प्रतीक्षा करा"),
    "status_translating": ("⏳ Translating... please wait", "⏳ भाषांतर करत आहे... कृपया प्रतीक्षा करा"),
    "search_no_results": ("No matches found", "काही जुळणी सापडली नाही"),
    "selected_member_prefix": ("Selected: ", "निवडलेले: "),
    "none_selected": ("No member selected", "कोणताही सेवक निवडलेला नाही"),
    "lbl_edit_search": ("Card Id to Edit:", "संपादनासाठी कार्ड आयडी:"),
    "btn_load_member": ("🔍 Load Sevak", "🔍 सेवक लोड करा"),
    "btn_update_member": ("💾 Update Sevak", "💾 सेवक अद्यतनित करा"),
    "no_member_found_title": ("Not Found", "सापडले नाही"),
    "no_member_found_body": ("No member found with that Card Id.", "या कार्ड आयडीसह कोणताही सेवक सापडला नाही."),
    "no_member_loaded_title": ("No Member Loaded", "सेवक लोड केलेला नाही"),
    "no_member_loaded_body": ("Please load a member by Card Id first.", "कृपया आधी कार्ड आयडीने सेवक लोड करा."),
    "update_success_body": ("Member updated successfully!", "सेवक यशस्वीरित्या अद्यतनित झाला!"),
    "nav_photo_list": ("Create Wari Photo List", "वारी फोटो यादी तयार करा"),
    "btn_generate_pdf": ("🖨 Generate PDF", "🖨 पीडीएफ तयार करा"),
    "pdf_no_attendees_title": ("No Attendees", "उपस्थित सेवक नाहीत"),
    "pdf_no_attendees_body": ("No attendees found for the selected Wari and year.",
                              "निवडलेल्या वारी आणि वर्षासाठी कोणतेही उपस्थित सेवक सापडले नाहीत."),
    "pdf_save_dialog_title": ("Save Wari Photo List PDF", "वारी फोटो यादी पीडीएफ जतन करा"),
    "pdf_success_body": ("PDF created successfully!", "पीडीएफ यशस्वीरित्या तयार झाली!"),
    "nav_pass_photo": ("Create Pass Photo PDF", "पास फोटो पीडीएफ तयार करा"),
    "btn_generate_pass_pdf": ("🖨 Generate Pass Photo PDF", "🖨 पास फोटो पीडीएफ तयार करा"),
    "pass_pdf_save_dialog_title": ("Save Pass Photo PDF", "पास फोटो पीडीएफ जतन करा"),
    "nav_import_records": ("Import Records", "रेकॉर्ड आयात करा"),
    "btn_select_excel": ("📂 Select Excel File", "📂 एक्सेल फाइल निवडा"),
    "lbl_select_sheet": ("Sheet:", "शीट:"),
    "chk_legacy_font": (
        "Marathi columns use AkrutiDynamicMar_BYogini font (auto-convert - best effort, review after import)",
        "मराठी स्तंभ AkrutiDynamicMar_BYogini फॉन्टमध्ये आहेत (स्वयं-रूपांतर - अंदाजे, आयातीनंतर तपासा)",
    ),
    "lbl_column_mapping": ("Map Excel columns to fields:", "एक्सेल स्तंभ फील्डशी जुळवा:"),
    "btn_preview_import": ("🔍 Preview", "🔍 पूर्वावलोकन"),
    "btn_import_all": ("📥 Import All Rows", "📥 सर्व नोंदी आयात करा"),
    "import_no_file_title": ("No File Selected", "फाइल निवडलेली नाही"),
    "import_no_file_body": ("Please select an Excel file first.", "कृपया आधी एक्सेल फाइल निवडा."),
    "import_no_rows_title": ("No Rows", "नोंदी नाहीत"),
    "import_no_rows_body": ("The selected sheet has no data rows.", "निवडलेल्या शीटमध्ये कोणतीही माहिती नाही."),
    "import_confirm_title": ("Confirm Import", "आयात निश्चित करा"),
    "import_no_card_id_title": ("Card Id Not Mapped", "कार्ड आयडी जुळलेला नाही"),
    "import_no_card_id_body": ("Please map a column to Card Id before importing.",
                               "आयात करण्यापूर्वी कृपया कार्ड आयडीशी एक स्तंभ जुळवा."),
    "(none)": ("(none)", "(काहीही नाही)"),
    "btn_bulk_translate": ("🌐 Translate All Missing Marathi + Link Photos",
                           "🌐 उर्वरित मराठी भाषांतर करा + फोटो जोडा"),
    "bulk_translate_confirm_title": ("Translate & Link Photos?", "भाषांतर आणि फोटो जोडायचे का?"),
    "bulk_translate_confirm_body": (
        "This scans every member in Sevak Details: fills in any blank "
        "Marathi field from its English value, and links any blank photo "
        "field to a Drive photo whose filename contains that member's "
        "Card Id. Existing Marathi text and photo links are left untouched. "
        "Continue?",
        "हे Sevak Details मधील प्रत्येक सेवक तपासून रिकाम्या मराठी फील्डमध्ये इंग्रजी "
        "मजकुरावरून भाषांतर भरेल, आणि ज्या सेवकाचा फोटो रिकामा आहे त्यासाठी कार्ड आयडी "
        "असलेल्या ड्राइव्ह फोटोची लिंक जोडेल. आधीपासून असलेला मराठी मजकूर आणि फोटो लिंक "
        "तसेच राहतील. पुढे जायचे का?",
    ),

}


DETAILS_SHEET_NAME = "Sevak Details"


def _open_sheet(worksheet_name):
    scopes = ["https://www.googleapis.com/auth/spreadsheets",
              "https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scopes)
    client = gspread.authorize(creds)
    return client.open_by_key(SHEET_ID).worksheet(worksheet_name)


def get_sheet(worksheet_name=DETAILS_SHEET_NAME):
    """Opens a specific worksheet (tab) by name within the spreadsheet.
    The workbook has multiple sheets, so callers must say which one they need.
    Retries on transient network errors (e.g. brief DNS hiccups)."""
    return with_retry(_open_sheet, worksheet_name)


# Maps our internal field keys to the actual column header text in the
# "Sevak Details" sheet. Columns are looked up by name (not position) since
# the real sheet has extra columns (mukkampost/taluka/district, both en and
# mr) interleaved between fields we do use - reading/writing by fixed
# position silently scrambles data whenever the sheet's layout differs from
# what the code assumes.
DETAILS_FIELDS = {
    "serial_no": "serial_no",
    "card_id": "card_id",
    "first_en": "first_name_en",
    "middle_en": "middle_name_en",
    "last_en": "last_name_en",
    "dob": "dob",
    "phone": "phone_number",
    "address_en": "address_en",
    "pincode": "pincode",
    "first_mr": "first_name_mr",
    "middle_mr": "middle_name_mr",
    "last_mr": "last_name_mr",
    "address_mr": "address_mr",
    "mukkam_en": "mukkam_en",
    "post_en": "post_en",
    "taluka_en": "taluka_en",
    "district_en": "district_en",
    "state_en": "state_en",
    "mukkam_mr": "mukkam_mr",
    "post_mr": "post_mr",
    "taluka_mr": "taluka_mr",
    "jilha_mr": "jilha_mr",
    "state_mr": "state_mr",
    "photo": "photo_path",
}


def get_header_map(sheet):
    """Returns {header_name: 0-based column index} for a sheet's first row."""
    header = sheet.row_values(1)
    return {name.strip(): i for i, name in enumerate(header) if name.strip()}


def row_to_details_record(row, header_map):
    """Converts a raw sheet row into a dict keyed by our internal field
    names, using the header map so column order doesn't matter."""
    def get(field_key):
        col_i = header_map.get(DETAILS_FIELDS[field_key])
        if col_i is None or col_i >= len(row):
            return ""
        return row[col_i].strip()
    return {key: get(key) for key in DETAILS_FIELDS}


def details_col_index(header_map, field_key):
    """1-based column index for a field key, or None if the sheet has no
    matching header (in which case that field is simply not written)."""
    col_i = header_map.get(DETAILS_FIELDS[field_key])
    return None if col_i is None else col_i + 1


def _get_text(field):
    return field.toPlainText() if isinstance(field, QTextEdit) else field.text()


def _set_text(field, text):
    if isinstance(field, QTextEdit):
        field.setPlainText(text)
    else:
        field.setText(text)


def _clear(field):
    field.clear()


def _to_number(value):
    """Converts a numeric-looking value (card_id, phone, pincode) to an int
    so it's written to the sheet as an actual number rather than text.
    Non-numeric or blank values are left untouched."""
    if isinstance(value, str):
        value = value.strip()
        if value.isdigit():
            return int(value)
    return value


class TranslateWorker(QThread):
    """Runs the (network-bound) translation calls off the UI thread, so a
    slow/blocked connection can't freeze - or appear to crash - the app."""
    finished_ok = pyqtSignal(dict)
    failed = pyqtSignal(str)

    def __init__(self, texts_by_index):
        super().__init__()
        self.texts_by_index = texts_by_index  # {pair_index: english_text}

    def run(self):
        results = {}
        try:
            for i, text in self.texts_by_index.items():
                results[i] = translate_to_marathi(text)
        except Exception as e:
            self.failed.emit(str(e))
            return
        self.finished_ok.emit(results)


class BulkTranslateWorker(QThread):
    """Scans the whole Sevak Details sheet for rows where an English field
    is filled in but its Marathi counterpart is blank, translates just
    those, and writes them all back in one batch update - so existing
    records typed in manually don't need to be opened one by one. Also
    fills in any blank photo link by matching a Drive photo whose filename
    contains the member's card_id."""
    progress = pyqtSignal(int, int)
    finished_ok = pyqtSignal(int, int)  # (translated_count, photo_linked_count)
    failed = pyqtSignal(str)

    EN_TO_MR_PAIRS = [
        ("first_en", "first_mr"), ("middle_en", "middle_mr"), ("last_en", "last_mr"),
        ("address_en", "address_mr"), ("mukkam_en", "mukkam_mr"), ("post_en", "post_mr"),
        ("taluka_en", "taluka_mr"), ("district_en", "jilha_mr"), ("state_en", "state_mr"),
    ]

    def _build_photo_link_by_card_id(self, card_id, drive_files):
        """Finds a Drive photo whose filename contains this card_id -
        possibly concatenated with other text (e.g. "525_ramesh.jpg",
        "IMG_525.jpg") - and returns its view link, or None if no match.
        Requires the card_id to appear as a whole number (not embedded
        inside a longer number, e.g. card_id "25" must not match "1250"),
        but allows any non-digit text immediately around it."""
        card_id = str(card_id).strip()
        if not card_id:
            return None
        pattern = re.compile(r"(?<!\d)" + re.escape(card_id) + r"(?!\d)")
        for f in drive_files:
            name_without_ext = os.path.splitext(f["name"])[0]
            if pattern.search(name_without_ext):
                return f"https://drive.google.com/file/d/{f['id']}/view"
        return None

    def run(self):
        try:
            sheet = get_sheet(DETAILS_SHEET_NAME)
            values = sheet.get_all_values()
            if not values or len(values) < 2:
                self.finished_ok.emit(0, 0)
                return

            try:
                drive_files = drive_helper.list_photos()
            except RuntimeError:
                drive_files = []  # photo linking is best-effort; translation still proceeds

            header_map = get_header_map(sheet)
            data_rows = values[1:]
            total = len(data_rows)
            batch_data = []
            translated_count = 0
            photo_linked_count = 0

            for i, row in enumerate(data_rows):
                self.progress.emit(i + 1, total)
                record = row_to_details_record(row, header_map)
                row_number = i + 2
                for en_key, mr_key in self.EN_TO_MR_PAIRS:
                    en_val = record.get(en_key, "")
                    mr_val = record.get(mr_key, "")
                    if en_val and not mr_val:
                        translated = translate_to_marathi(en_val)
                        if translated:
                            col = details_col_index(header_map, mr_key)
                            if col is not None:
                                batch_data.append({
                                    "range": gspread.utils.rowcol_to_a1(row_number, col),
                                    "values": [[translated]],
                                })
                                translated_count += 1

                if not record.get("photo") and drive_files:
                    link = self._build_photo_link_by_card_id(record.get("card_id", ""), drive_files)
                    if link:
                        col = details_col_index(header_map, "photo")
                        if col is not None:
                            batch_data.append({
                                "range": gspread.utils.rowcol_to_a1(row_number, col),
                                "values": [[link]],
                            })
                            photo_linked_count += 1

            if batch_data:
                sheet.batch_update(batch_data)
            self.finished_ok.emit(translated_count, photo_linked_count)
        except Exception as e:
            self.failed.emit(str(e))


class SevakJodaForm(QMainWindow):
    def __init__(self):
        super().__init__()
        self.lang = "en"
        self.photo_path = None
        self.photo_drive_id = None
        self.photo_drive_link = None
        self.edit_photo_path = None
        self.edit_photo_drive_id = None
        self.edit_photo_drive_link = None
        self.edit_original_photo_value = None
        self.edit_row_number = None
        self._busy_depth = 0
        self.import_workbook_path = None
        self.import_header = []
        self.import_data_rows = []
        self.resize(1100, 700)

        central = QWidget()
        self.setCentralWidget(central)
        root_layout = QHBoxLayout()
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)
        central.setLayout(root_layout)

        self.selected_wari_member = None

        self.stack = QStackedWidget()
        self.add_member_page = self._build_add_member_page()
        self.edit_member_page = self._build_edit_member_page()
        self.vari_member_page = self._build_wari_attendees_page()
        self.photo_list_page = self._build_photo_list_page()
        self.pass_photo_page = self._build_pass_photo_page()
        self.import_records_page = self._build_import_records_page()
        self.stack.addWidget(self.add_member_page)
        self.stack.addWidget(self.edit_member_page)
        self.stack.addWidget(self.vari_member_page)
        self.stack.addWidget(self.photo_list_page)
        self.stack.addWidget(self.pass_photo_page)
        self.stack.addWidget(self.import_records_page)
        self.stack.currentChanged.connect(self.on_page_changed)

        root_layout.addWidget(self._build_sidebar())
        root_layout.addWidget(self.stack, 1)

        self.apply_language()

    def on_page_changed(self, _index):
        if self.stack.currentWidget() is self.vari_member_page:
            self.refresh_wari_attendees_list()

    # ---------- Busy/loading state ----------

    def _set_busy(self, busy, message=None):
        """Locks the whole window (so no button/field anywhere can be
        clicked) and shows a wait cursor + status bar message while any
        blocking operation (network call, scan, PDF generation, etc.) runs.
        Reentrant: safe to call from a busy-wrapped method that itself
        calls another busy-wrapped method - the window only re-enables
        once every nested call has finished."""
        idx = 0 if self.lang == "en" else 1
        if busy:
            self._busy_depth += 1
            if self._busy_depth == 1:
                self.centralWidget().setEnabled(False)
            QApplication.setOverrideCursor(Qt.WaitCursor)
            self.statusBar().showMessage(message or LABELS["status_processing"][idx])
        else:
            # guard against being called more times than _set_busy(True) was
            # (harmless no-op instead of underflowing Qt's cursor stack)
            if self._busy_depth > 0:
                QApplication.restoreOverrideCursor()
                self._busy_depth -= 1
            if self._busy_depth == 0:
                self.centralWidget().setEnabled(True)
                self.statusBar().clearMessage()
        QApplication.processEvents()

    # ---------- Sidebar (navigation + sheet link + language) ----------

    def _build_sidebar(self):
        self.sidebar = QFrame()
        self.sidebar.setFixedWidth(SIDEBAR_WIDTH_EXPANDED)
        self.sidebar.setStyleSheet("QFrame { background-color: #2c3e50; }")
        self.sidebar_collapsed = False

        layout = QVBoxLayout()
        layout.setContentsMargins(6, 6, 6, 6)
        self.sidebar.setLayout(layout)

        toggle_btn = QPushButton("☰")
        toggle_btn.setFixedHeight(32)
        toggle_btn.clicked.connect(self.toggle_sidebar)
        layout.addWidget(toggle_btn)

        self.nav_buttons = {}
        nav_defs = [
            ("nav_add_member", self.add_member_page),
            ("nav_edit_member", self.edit_member_page),
            ("nav_vari_member", self.vari_member_page),
            ("nav_photo_list", self.photo_list_page),
            ("nav_pass_photo", self.pass_photo_page),
            ("nav_import_records", self.import_records_page),
        ]
        for key, page in nav_defs:
            btn = QPushButton()
            btn.setStyleSheet(
                "QPushButton { text-align: left; padding: 8px; color: white; "
                "background-color: #34495e; border: none; border-radius: 4px; }"
                "QPushButton:hover { background-color: #3d566e; }"
            )
            btn.clicked.connect(lambda _, p=page: self.stack.setCurrentWidget(p))
            layout.addWidget(btn)
            self.nav_buttons[key] = btn

        self.sheet_link_btn = QPushButton()
        self.sheet_link_btn.setStyleSheet(
            "QPushButton { text-align: left; padding: 8px; color: white; "
            "background-color: #34495e; border: none; border-radius: 4px; }"
            "QPushButton:hover { background-color: #3d566e; }"
        )
        self.sheet_link_btn.clicked.connect(self.open_sheet_link)
        layout.addWidget(self.sheet_link_btn)

        self.drive_link_btn = QPushButton()
        self.drive_link_btn.setStyleSheet(
            "QPushButton { text-align: left; padding: 8px; color: white; "
            "background-color: #34495e; border: none; border-radius: 4px; }"
            "QPushButton:hover { background-color: #3d566e; }"
        )
        self.drive_link_btn.clicked.connect(self.open_drive_link)
        layout.addWidget(self.drive_link_btn)

        layout.addStretch()

        self.lang_btn = QPushButton()
        self.lang_btn.setStyleSheet(
            "QPushButton { padding: 8px; color: white; background-color: #16a085; "
            "border: none; border-radius: 4px; }"
        )
        self.lang_btn.clicked.connect(self.toggle_language)
        layout.addWidget(self.lang_btn)

        return self.sidebar

    def toggle_sidebar(self):
        self.sidebar_collapsed = not self.sidebar_collapsed
        if self.sidebar_collapsed:
            self.sidebar.setFixedWidth(SIDEBAR_WIDTH_COLLAPSED)
            for btn in self.nav_buttons.values():
                btn.setText("")
            self.sheet_link_btn.setText("")
            self.drive_link_btn.setText("")
            self.lang_btn.setText("")
        else:
            self.sidebar.setFixedWidth(SIDEBAR_WIDTH_EXPANDED)
            self._refresh_nav_labels()

    def _refresh_nav_labels(self):
        idx = 0 if self.lang == "en" else 1
        for key, btn in self.nav_buttons.items():
            btn.setText("" if self.sidebar_collapsed else LABELS[key][idx])
        self.sheet_link_btn.setText("" if self.sidebar_collapsed else LABELS["open_sheet"][idx])
        self.drive_link_btn.setText("" if self.sidebar_collapsed else LABELS["open_drive"][idx])
        self.lang_btn.setText("" if self.sidebar_collapsed else LABELS["lang_toggle"][idx])

    def open_sheet_link(self):
        QDesktopServices.openUrl(QUrl(SHEET_URL))

    def open_drive_link(self):
        QDesktopServices.openUrl(QUrl(DRIVE_FOLDER_URL))

    # ---------- Language ----------

    def toggle_language(self):
        self.lang = "mr" if self.lang == "en" else "en"
        self.apply_language()

    def apply_language(self):
        idx = 0 if self.lang == "en" else 1
        t = lambda key: LABELS[key][idx]

        self.setWindowTitle(t("window_title"))
        self._refresh_nav_labels()

        for key, label_widget in self.form_labels.items():
            label_widget.setText(t(key))

        self.translate_btn.setText(t("btn_translate"))
        self.capture_btn.setText(t("btn_scan"))
        self.browse_btn.setText(t("btn_browse"))
        self.drive_pick_btn.setText(t("btn_drive_pick"))
        self.save_btn.setText(t("btn_save"))
        if self.photo_path is None:
            self.photo_preview.setText(t("no_photo"))

        for key, label_widget in self.edit_form_labels.items():
            label_widget.setText(t(key))
        self.edit_search_label.setText(t("lbl_edit_search"))
        self.edit_load_btn.setText(t("btn_load_member"))
        self.edit_translate_btn.setText(t("btn_translate"))
        self.edit_capture_btn.setText(t("btn_scan"))
        self.edit_browse_btn.setText(t("btn_browse"))
        self.edit_drive_pick_btn.setText(t("btn_drive_pick"))
        self.edit_update_btn.setText(t("btn_update_member"))
        if self.edit_photo_path is None:
            self.edit_photo_preview.setText(t("no_photo"))

        self._refresh_wari_combo_labels()
        self.wari_select_label.setText(t("lbl_wari_select"))
        self.wari_year_label.setText(t("lbl_wari_year"))
        self.wari_search_label.setText(t("lbl_search_member"))
        self.wari_search_btn.setText(t("btn_search"))
        self.wari_add_btn.setText(t("btn_add_to_wari"))
        self.wari_show_list_btn.setText(t("btn_show_list"))
        self.wari_remove_btn.setText(t("btn_remove_attendee"))
        self.wari_attendees_label.setText(t("lbl_attendees_list"))
        if self.selected_wari_member is None:
            self.wari_selected_label.setText(t("none_selected"))

        self._refresh_photo_list_wari_combo_labels()
        self.photo_list_select_label.setText(t("lbl_wari_select"))
        self.photo_list_year_label.setText(t("lbl_wari_year"))
        self.photo_list_generate_btn.setText(t("btn_generate_pdf"))

        self._refresh_pass_photo_wari_combo_labels()
        self.pass_photo_select_label.setText(t("lbl_wari_select"))
        self.pass_photo_year_label.setText(t("lbl_wari_year"))
        self.pass_photo_generate_btn.setText(t("btn_generate_pass_pdf"))

        self.import_select_file_btn.setText(t("btn_select_excel"))
        self.import_sheet_label.setText(t("lbl_select_sheet"))
        self.import_legacy_font_checkbox.setText(t("chk_legacy_font"))
        self.import_mapping_label.setText(t("lbl_column_mapping"))
        self.import_preview_btn.setText(t("btn_preview_import"))
        self.import_all_btn.setText(t("btn_import_all"))
        self.bulk_translate_btn.setText(t("btn_bulk_translate"))
        self._refresh_import_field_labels()

    # ---------- Add Member page ----------

    def _build_add_member_page(self):
        page = QWidget()
        outer_layout = QVBoxLayout()
        page.setLayout(outer_layout)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_content = QWidget()
        scroll.setWidget(scroll_content)
        main_layout = QVBoxLayout()
        scroll_content.setLayout(main_layout)
        outer_layout.addWidget(scroll)

        # English fields on the left, Marathi/other fields on the right,
        # so the form doesn't overflow vertically off the screen.
        form_en = QFormLayout()
        form_mr = QFormLayout()
        self.form_labels = {}

        def add_row(key, field, form=form_en):
            label = QLabel()
            self.form_labels[key] = label
            form.addRow(label, field)

        # English fields
        eng_font = QFont(ENGLISH_FONT, 11)
        self.card_id = QLineEdit(); self.card_id.setFont(eng_font)
        self.first_name_en = QLineEdit(); self.first_name_en.setFont(eng_font)
        self.middle_name_en = QLineEdit(); self.middle_name_en.setFont(eng_font)
        self.last_name_en = QLineEdit(); self.last_name_en.setFont(eng_font)
        self.dob = QLineEdit(); self.dob.setFont(eng_font)
        self.phone_number = QLineEdit(); self.phone_number.setFont(eng_font)
        self.address_en = QTextEdit(); self.address_en.setFont(eng_font)
        self.address_en.setMinimumHeight(90)
        self.mukkam_en = QLineEdit(); self.mukkam_en.setFont(eng_font)
        self.post_en = QLineEdit(); self.post_en.setFont(eng_font)
        self.taluka_en = QLineEdit(); self.taluka_en.setFont(eng_font)
        self.district_en = QLineEdit(); self.district_en.setFont(eng_font)
        self.state_en = QLineEdit(); self.state_en.setFont(eng_font)
        self.pincode = QLineEdit(); self.pincode.setFont(eng_font)

        add_row("lbl_card_id", self.card_id)
        add_row("lbl_first_en", self.first_name_en)
        add_row("lbl_middle_en", self.middle_name_en)
        add_row("lbl_last_en", self.last_name_en)
        add_row("lbl_dob", self.dob)
        add_row("lbl_phone", self.phone_number)
        add_row("lbl_address_en", self.address_en)
        add_row("lbl_mukkam_en", self.mukkam_en)
        add_row("lbl_post_en", self.post_en)
        add_row("lbl_taluka_en", self.taluka_en)
        add_row("lbl_district_en", self.district_en)
        add_row("lbl_state_en", self.state_en)
        add_row("lbl_pincode", self.pincode)

        # Marathi fields
        mr_font = QFont(MARATHI_FONT, 12)
        self.first_name_mr = QLineEdit(); self.first_name_mr.setFont(mr_font)
        self.middle_name_mr = QLineEdit(); self.middle_name_mr.setFont(mr_font)
        self.last_name_mr = QLineEdit(); self.last_name_mr.setFont(mr_font)
        self.address_mr = QTextEdit(); self.address_mr.setFont(mr_font)
        self.address_mr.setMinimumHeight(90)
        self.mukkam_mr = QLineEdit(); self.mukkam_mr.setFont(mr_font)
        self.post_mr = QLineEdit(); self.post_mr.setFont(mr_font)
        self.taluka_mr = QLineEdit(); self.taluka_mr.setFont(mr_font)
        self.jilha_mr = QLineEdit(); self.jilha_mr.setFont(mr_font)
        self.state_mr = QLineEdit(); self.state_mr.setFont(mr_font)

        self.marathi_keyboard = MarathiKeyboard(self)

        # pairs of (english field, marathi field) kept in sync by translation
        self.translation_pairs = [
            (self.first_name_en, self.first_name_mr),
            (self.middle_name_en, self.middle_name_mr),
            (self.last_name_en, self.last_name_mr),
            (self.address_en, self.address_mr),
            (self.mukkam_en, self.mukkam_mr),
            (self.post_en, self.post_mr),
            (self.taluka_en, self.taluka_mr),
            (self.district_en, self.jilha_mr),
            (self.state_en, self.state_mr),
        ]

        add_row("lbl_first_mr", self._mr_row(self.first_name_mr), form_mr)
        add_row("lbl_middle_mr", self._mr_row(self.middle_name_mr), form_mr)
        add_row("lbl_last_mr", self._mr_row(self.last_name_mr), form_mr)
        add_row("lbl_address_mr", self._mr_row(self.address_mr), form_mr)
        add_row("lbl_mukkam_mr", self._mr_row(self.mukkam_mr), form_mr)
        add_row("lbl_post_mr", self._mr_row(self.post_mr), form_mr)
        add_row("lbl_taluka_mr", self._mr_row(self.taluka_mr), form_mr)
        add_row("lbl_jilha_mr", self._mr_row(self.jilha_mr), form_mr)
        add_row("lbl_state_mr", self._mr_row(self.state_mr), form_mr)

        columns_layout = QHBoxLayout()
        left_container = QWidget()
        left_container.setLayout(form_en)
        right_container = QWidget()
        right_container.setLayout(form_mr)
        columns_layout.addWidget(left_container)
        columns_layout.addWidget(right_container)
        main_layout.addLayout(columns_layout)

        self.translate_btn = QPushButton()
        self.translate_btn.clicked.connect(lambda: self.translate_fields())
        main_layout.addWidget(self.translate_btn)

        # Photo section
        photo_layout = QHBoxLayout()
        self.photo_preview = QLabel()
        self.photo_preview.setFixedSize(150, 190)
        self.photo_preview.setStyleSheet("border: 1px solid gray;")
        self.photo_preview.setAlignment(Qt.AlignCenter)

        self.capture_btn = QPushButton()
        self.capture_btn.clicked.connect(lambda: self.run_scan(context="add"))

        self.browse_btn = QPushButton()
        self.browse_btn.clicked.connect(lambda: self.browse_local_photo(context="add"))

        self.drive_pick_btn = QPushButton()
        self.drive_pick_btn.clicked.connect(lambda: self.browse_drive_photo(context="add"))

        photo_btn_layout = QVBoxLayout()
        photo_btn_layout.addWidget(self.capture_btn)
        photo_btn_layout.addWidget(self.browse_btn)
        photo_btn_layout.addWidget(self.drive_pick_btn)

        photo_layout.addWidget(self.photo_preview)
        photo_layout.addLayout(photo_btn_layout)
        main_layout.addLayout(photo_layout)

        # Save button
        self.save_btn = QPushButton()
        self.save_btn.setStyleSheet("font-size: 14px; padding: 10px; background-color: #4CAF50; color: white;")
        self.save_btn.clicked.connect(self.save_member)
        main_layout.addWidget(self.save_btn)

        return page

    # ---------- Edit Sevak page ----------

    def _build_edit_member_page(self):
        page = QWidget()
        outer_layout = QVBoxLayout()
        page.setLayout(outer_layout)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_content = QWidget()
        scroll.setWidget(scroll_content)
        main_layout = QVBoxLayout()
        scroll_content.setLayout(main_layout)

        search_row = QHBoxLayout()
        self.edit_search_label = QLabel()
        self.edit_card_id_search = QLineEdit()
        self.edit_load_btn = QPushButton()
        self.edit_load_btn.clicked.connect(self.load_edit_member)
        self.edit_card_id_search.returnPressed.connect(self.load_edit_member)
        search_row.addWidget(self.edit_search_label)
        search_row.addWidget(self.edit_card_id_search)
        search_row.addWidget(self.edit_load_btn)
        outer_layout.addLayout(search_row)
        outer_layout.addWidget(scroll)

        # English fields on the left, Marathi/other fields on the right,
        # so the form doesn't overflow vertically off the screen.
        form_en = QFormLayout()
        form_mr = QFormLayout()
        self.edit_form_labels = {}

        def add_row(key, field, form=form_en):
            label = QLabel()
            self.edit_form_labels[key] = label
            form.addRow(label, field)

        eng_font = QFont(ENGLISH_FONT, 11)
        self.edit_card_id = QLineEdit(); self.edit_card_id.setFont(eng_font)
        self.edit_first_name_en = QLineEdit(); self.edit_first_name_en.setFont(eng_font)
        self.edit_middle_name_en = QLineEdit(); self.edit_middle_name_en.setFont(eng_font)
        self.edit_last_name_en = QLineEdit(); self.edit_last_name_en.setFont(eng_font)
        self.edit_dob = QLineEdit(); self.edit_dob.setFont(eng_font)
        self.edit_phone_number = QLineEdit(); self.edit_phone_number.setFont(eng_font)
        self.edit_address_en = QTextEdit(); self.edit_address_en.setFont(eng_font)
        self.edit_address_en.setMinimumHeight(90)
        self.edit_mukkam_en = QLineEdit(); self.edit_mukkam_en.setFont(eng_font)
        self.edit_post_en = QLineEdit(); self.edit_post_en.setFont(eng_font)
        self.edit_taluka_en = QLineEdit(); self.edit_taluka_en.setFont(eng_font)
        self.edit_district_en = QLineEdit(); self.edit_district_en.setFont(eng_font)
        self.edit_state_en = QLineEdit(); self.edit_state_en.setFont(eng_font)
        self.edit_pincode = QLineEdit(); self.edit_pincode.setFont(eng_font)

        add_row("lbl_card_id", self.edit_card_id)
        add_row("lbl_first_en", self.edit_first_name_en)
        add_row("lbl_middle_en", self.edit_middle_name_en)
        add_row("lbl_last_en", self.edit_last_name_en)
        add_row("lbl_dob", self.edit_dob)
        add_row("lbl_phone", self.edit_phone_number)
        add_row("lbl_address_en", self.edit_address_en)
        add_row("lbl_mukkam_en", self.edit_mukkam_en)
        add_row("lbl_post_en", self.edit_post_en)
        add_row("lbl_taluka_en", self.edit_taluka_en)
        add_row("lbl_district_en", self.edit_district_en)
        add_row("lbl_state_en", self.edit_state_en)
        add_row("lbl_pincode", self.edit_pincode)

        mr_font = QFont(MARATHI_FONT, 12)
        self.edit_first_name_mr = QLineEdit(); self.edit_first_name_mr.setFont(mr_font)
        self.edit_middle_name_mr = QLineEdit(); self.edit_middle_name_mr.setFont(mr_font)
        self.edit_last_name_mr = QLineEdit(); self.edit_last_name_mr.setFont(mr_font)
        self.edit_address_mr = QTextEdit(); self.edit_address_mr.setFont(mr_font)
        self.edit_address_mr.setMinimumHeight(90)
        self.edit_mukkam_mr = QLineEdit(); self.edit_mukkam_mr.setFont(mr_font)
        self.edit_post_mr = QLineEdit(); self.edit_post_mr.setFont(mr_font)
        self.edit_taluka_mr = QLineEdit(); self.edit_taluka_mr.setFont(mr_font)
        self.edit_jilha_mr = QLineEdit(); self.edit_jilha_mr.setFont(mr_font)
        self.edit_state_mr = QLineEdit(); self.edit_state_mr.setFont(mr_font)

        self.edit_translation_pairs = [
            (self.edit_first_name_en, self.edit_first_name_mr),
            (self.edit_middle_name_en, self.edit_middle_name_mr),
            (self.edit_last_name_en, self.edit_last_name_mr),
            (self.edit_address_en, self.edit_address_mr),
            (self.edit_mukkam_en, self.edit_mukkam_mr),
            (self.edit_post_en, self.edit_post_mr),
            (self.edit_taluka_en, self.edit_taluka_mr),
            (self.edit_district_en, self.edit_jilha_mr),
            (self.edit_state_en, self.edit_state_mr),
        ]

        add_row("lbl_first_mr", self._mr_row(self.edit_first_name_mr), form_mr)
        add_row("lbl_middle_mr", self._mr_row(self.edit_middle_name_mr), form_mr)
        add_row("lbl_last_mr", self._mr_row(self.edit_last_name_mr), form_mr)
        add_row("lbl_address_mr", self._mr_row(self.edit_address_mr), form_mr)
        add_row("lbl_mukkam_mr", self._mr_row(self.edit_mukkam_mr), form_mr)
        add_row("lbl_post_mr", self._mr_row(self.edit_post_mr), form_mr)
        add_row("lbl_taluka_mr", self._mr_row(self.edit_taluka_mr), form_mr)
        add_row("lbl_jilha_mr", self._mr_row(self.edit_jilha_mr), form_mr)
        add_row("lbl_state_mr", self._mr_row(self.edit_state_mr), form_mr)

        columns_layout = QHBoxLayout()
        left_container = QWidget()
        left_container.setLayout(form_en)
        right_container = QWidget()
        right_container.setLayout(form_mr)
        columns_layout.addWidget(left_container)
        columns_layout.addWidget(right_container)
        main_layout.addLayout(columns_layout)

        self.edit_translate_btn = QPushButton()
        self.edit_translate_btn.clicked.connect(lambda: self.translate_fields(self.edit_translation_pairs))
        main_layout.addWidget(self.edit_translate_btn)

        photo_layout = QHBoxLayout()
        self.edit_photo_preview = QLabel()
        self.edit_photo_preview.setFixedSize(150, 190)
        self.edit_photo_preview.setStyleSheet("border: 1px solid gray;")
        self.edit_photo_preview.setAlignment(Qt.AlignCenter)

        self.edit_capture_btn = QPushButton()
        self.edit_capture_btn.clicked.connect(lambda: self.run_scan(context="edit"))

        self.edit_browse_btn = QPushButton()
        self.edit_browse_btn.clicked.connect(lambda: self.browse_local_photo(context="edit"))

        self.edit_drive_pick_btn = QPushButton()
        self.edit_drive_pick_btn.clicked.connect(lambda: self.browse_drive_photo(context="edit"))

        photo_btn_layout = QVBoxLayout()
        photo_btn_layout.addWidget(self.edit_capture_btn)
        photo_btn_layout.addWidget(self.edit_browse_btn)
        photo_btn_layout.addWidget(self.edit_drive_pick_btn)

        photo_layout.addWidget(self.edit_photo_preview)
        photo_layout.addLayout(photo_btn_layout)
        main_layout.addLayout(photo_layout)

        self.edit_update_btn = QPushButton()
        self.edit_update_btn.setStyleSheet("font-size: 14px; padding: 10px; background-color: #4CAF50; color: white;")
        self.edit_update_btn.clicked.connect(self.update_member)
        main_layout.addWidget(self.edit_update_btn)

        return page

    # ---------- Wari Attendees page ----------

    def _build_wari_attendees_page(self):
        page = QWidget()
        layout = QVBoxLayout()
        page.setLayout(layout)

        selector_form = QFormLayout()

        self.wari_combo = QComboBox()
        for en_name, mr_name in WARI_OPTIONS:
            self.wari_combo.addItem(en_name, en_name)
        self.wari_combo.currentIndexChanged.connect(self.refresh_wari_attendees_list)
        self.wari_select_label = QLabel()
        selector_form.addRow(self.wari_select_label, self.wari_combo)

        self.wari_year_spin = QSpinBox()
        self.wari_year_spin.setRange(2000, 2100)
        self.wari_year_spin.setValue(datetime.now().year)
        self.wari_year_spin.valueChanged.connect(self.refresh_wari_attendees_list)
        self.wari_year_label = QLabel()
        selector_form.addRow(self.wari_year_label, self.wari_year_spin)

        layout.addLayout(selector_form)

        search_row = QHBoxLayout()
        self.wari_search_label = QLabel()
        self.wari_search_input = QLineEdit()
        self.wari_search_btn = QPushButton()
        self.wari_search_btn.clicked.connect(self.search_wari_members)
        self.wari_search_input.returnPressed.connect(self.search_wari_members)
        search_row.addWidget(self.wari_search_input)
        search_row.addWidget(self.wari_search_btn)
        layout.addWidget(self.wari_search_label)
        layout.addLayout(search_row)

        self.wari_search_results = QListWidget()
        self.wari_search_results.setFixedHeight(120)
        self.wari_search_results.itemClicked.connect(self.select_wari_member)
        layout.addWidget(self.wari_search_results)

        self.wari_selected_label = QLabel()
        layout.addWidget(self.wari_selected_label)

        add_show_row = QHBoxLayout()
        self.wari_add_btn = QPushButton()
        self.wari_add_btn.clicked.connect(self.save_wari_attendee)
        self.wari_show_list_btn = QPushButton()
        self.wari_show_list_btn.clicked.connect(lambda: self.refresh_wari_attendees_list())
        add_show_row.addWidget(self.wari_add_btn)
        add_show_row.addWidget(self.wari_show_list_btn)
        layout.addLayout(add_show_row)

        self.wari_attendees_label = QLabel()
        layout.addWidget(self.wari_attendees_label)

        self.wari_attendees_listbox = QListWidget()
        layout.addWidget(self.wari_attendees_listbox, 1)

        self.wari_remove_btn = QPushButton()
        self.wari_remove_btn.clicked.connect(self.remove_wari_attendee)
        layout.addWidget(self.wari_remove_btn)

        return page

    def _refresh_wari_combo_labels(self):
        idx = 0 if self.lang == "en" else 1
        for i, (en_name, mr_name) in enumerate(WARI_OPTIONS):
            self.wari_combo.setItemText(i, en_name if idx == 0 else mr_name)

    def fetch_details_records(self):
        """Reads every member from the Details sheet for name/card-id search."""
        sheet = get_sheet(DETAILS_SHEET_NAME)
        values = sheet.get_all_values()
        if not values:
            return []
        header_map = get_header_map(sheet)
        records = []
        for row in values[1:]:
            record = row_to_details_record(row, header_map)
            if not record["card_id"]:
                continue
            records.append(record)
        return records

    @staticmethod
    def _full_name_en(record):
        return " ".join(filter(None, [record["first_en"], record["middle_en"], record["last_en"]]))

    def search_wari_members(self):
        idx = 0 if self.lang == "en" else 1
        term = self.wari_search_input.text().strip().lower()
        self.wari_search_results.clear()
        if not term:
            return

        self._set_busy(True, LABELS["status_searching"][idx])
        try:
            records = self.fetch_details_records()
        except Exception as e:
            self._set_busy(False)
            QMessageBox.critical(self, LABELS["error_title"][idx], f"Search failed: {e}")
            return
        self._set_busy(False)

        matches = [
            r for r in records
            if term in r["card_id"].lower() or term in self._full_name_en(r).lower()
        ]
        if not matches:
            item = QListWidgetItem(LABELS["search_no_results"][idx])
            item.setFlags(Qt.NoItemFlags)
            self.wari_search_results.addItem(item)
            return

        for r in matches:
            item = QListWidgetItem(f'{r["card_id"]} — {self._full_name_en(r)}')
            item.setData(Qt.UserRole, r)
            self.wari_search_results.addItem(item)

    def select_wari_member(self, item):
        record = item.data(Qt.UserRole)
        if not record:
            return
        idx = 0 if self.lang == "en" else 1
        self.selected_wari_member = record
        self.wari_selected_label.setText(
            f'{LABELS["selected_member_prefix"][idx]}{record["card_id"]} — {self._full_name_en(record)}'
        )

    @staticmethod
    def _ensure_wari_header(sheet):
        if not sheet.get_all_values():
            sheet.append_row(["Wari Name", "Wari Year", "Card Id", "Name English"])

    @staticmethod
    def _find_wari_insert_position(values, wari_name, wari_year, new_last_name):
        """Returns the 1-based sheet row index to insert a new attendee at,
        keeping attendees of the same Wari+year sorted by last name (the
        last word of the stored "Name English" column)."""
        new_last_name = (new_last_name or "").strip().lower()
        last_matching_row = None
        for i, row in enumerate(values[1:], start=2):
            if len(row) < 4 or row[0] != wari_name or str(row[1]) != str(wari_year):
                continue
            last_matching_row = i
            existing_name = row[3].strip()
            existing_last_name = existing_name.split()[-1].lower() if existing_name else ""
            if new_last_name < existing_last_name:
                return i
        return (last_matching_row + 1) if last_matching_row is not None else len(values) + 1

    def save_wari_attendee(self):
        idx = 0 if self.lang == "en" else 1
        if not self.selected_wari_member:
            QMessageBox.warning(self, LABELS["no_member_selected_title"][idx],
                                 LABELS["no_member_selected_body"][idx])
            return

        wari_name = self.wari_combo.currentData()
        wari_year = self.wari_year_spin.value()
        card_id = self.selected_wari_member["card_id"]
        full_name = self._full_name_en(self.selected_wari_member)
        last_name = self.selected_wari_member.get("last_en", "")

        self._set_busy(True, LABELS["status_saving"][idx])
        try:
            sheet = get_sheet(WARI_SHEET_NAME)
            self._ensure_wari_header(sheet)
            values = sheet.get_all_values()
            duplicate = any(
                len(row) >= 3 and row[0] == wari_name and str(row[1]) == str(wari_year) and row[2] == card_id
                for row in values[1:]
            )
            if not duplicate:
                insert_at = self._find_wari_insert_position(values, wari_name, wari_year, last_name)
                sheet.insert_row([wari_name, wari_year, _to_number(card_id), full_name], index=insert_at)
        except Exception as e:
            self._set_busy(False)
            QMessageBox.critical(self, LABELS["error_title"][idx], f"Failed to save: {e}")
            return
        self._set_busy(False)

        if duplicate:
            QMessageBox.information(self, LABELS["duplicate_title"][idx], LABELS["duplicate_body"][idx])
            return
        QMessageBox.information(self, LABELS["success_title"][idx], LABELS["wari_added_body"][idx])
        self.refresh_wari_attendees_list()

    def refresh_wari_attendees_list(self):
        idx = 0 if self.lang == "en" else 1
        self.wari_attendees_listbox.clear()
        self._set_busy(True, LABELS["status_loading"][idx])
        try:
            sheet = get_sheet(WARI_SHEET_NAME)
            values = sheet.get_all_values()
        except Exception:
            return
        finally:
            self._set_busy(False)

        wari_name = self.wari_combo.currentData()
        wari_year = str(self.wari_year_spin.value())
        for row_number, row in enumerate(values[1:], start=2):
            if len(row) >= 4 and row[0] == wari_name and str(row[1]) == wari_year:
                item = QListWidgetItem(f'{row[2]} — {row[3]}')
                item.setData(Qt.UserRole, row_number)
                self.wari_attendees_listbox.addItem(item)

    def remove_wari_attendee(self):
        idx = 0 if self.lang == "en" else 1
        item = self.wari_attendees_listbox.currentItem()
        if not item:
            QMessageBox.warning(self, LABELS["no_attendee_selected_title"][idx],
                                 LABELS["no_attendee_selected_body"][idx])
            return

        confirm = QMessageBox.question(
            self, LABELS["remove_confirm_title"][idx], LABELS["remove_confirm_body"][idx],
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
        )
        if confirm != QMessageBox.Yes:
            return

        row_number = item.data(Qt.UserRole)
        self._set_busy(True, LABELS["status_removing"][idx])
        try:
            sheet = get_sheet(WARI_SHEET_NAME)
            sheet.delete_rows(row_number)
        except Exception as e:
            self._set_busy(False)
            QMessageBox.critical(self, LABELS["error_title"][idx], f"Failed to remove: {e}")
            return
        self._set_busy(False)
        self.refresh_wari_attendees_list()
        QMessageBox.information(self, LABELS["success_title"][idx], LABELS["remove_success_body"][idx])

    # ---------- Create Wari Photo List page ----------

    def _build_photo_list_page(self):
        page = QWidget()
        layout = QVBoxLayout()
        page.setLayout(layout)

        selector_form = QFormLayout()

        self.photo_list_wari_combo = QComboBox()
        for en_name, mr_name in WARI_OPTIONS:
            self.photo_list_wari_combo.addItem(en_name, en_name)
        self.photo_list_select_label = QLabel()
        selector_form.addRow(self.photo_list_select_label, self.photo_list_wari_combo)

        self.photo_list_year_spin = QSpinBox()
        self.photo_list_year_spin.setRange(2000, 2100)
        self.photo_list_year_spin.setValue(datetime.now().year)
        self.photo_list_year_label = QLabel()
        selector_form.addRow(self.photo_list_year_label, self.photo_list_year_spin)

        layout.addLayout(selector_form)

        self.photo_list_generate_btn = QPushButton()
        self.photo_list_generate_btn.setStyleSheet(
            "font-size: 14px; padding: 10px; background-color: #4CAF50; color: white;"
        )
        self.photo_list_generate_btn.clicked.connect(self.generate_wari_photo_list_pdf)
        layout.addWidget(self.photo_list_generate_btn)

        layout.addStretch()

        return page

    def _refresh_photo_list_wari_combo_labels(self):
        idx = 0 if self.lang == "en" else 1
        for i, (en_name, mr_name) in enumerate(WARI_OPTIONS):
            self.photo_list_wari_combo.setItemText(i, en_name if idx == 0 else mr_name)

    # ---------- Create Pass Photo PDF page ----------

    def _build_pass_photo_page(self):
        page = QWidget()
        layout = QVBoxLayout()
        page.setLayout(layout)

        selector_form = QFormLayout()

        self.pass_photo_wari_combo = QComboBox()
        for en_name, mr_name in WARI_OPTIONS:
            self.pass_photo_wari_combo.addItem(en_name, en_name)
        self.pass_photo_select_label = QLabel()
        selector_form.addRow(self.pass_photo_select_label, self.pass_photo_wari_combo)

        self.pass_photo_year_spin = QSpinBox()
        self.pass_photo_year_spin.setRange(2000, 2100)
        self.pass_photo_year_spin.setValue(datetime.now().year)
        self.pass_photo_year_label = QLabel()
        selector_form.addRow(self.pass_photo_year_label, self.pass_photo_year_spin)

        layout.addLayout(selector_form)

        self.pass_photo_generate_btn = QPushButton()
        self.pass_photo_generate_btn.setStyleSheet(
            "font-size: 14px; padding: 10px; background-color: #4CAF50; color: white;"
        )
        self.pass_photo_generate_btn.clicked.connect(self.generate_pass_photo_pdf)
        layout.addWidget(self.pass_photo_generate_btn)

        layout.addStretch()

        return page

    def _refresh_pass_photo_wari_combo_labels(self):
        idx = 0 if self.lang == "en" else 1
        for i, (en_name, mr_name) in enumerate(WARI_OPTIONS):
            self.pass_photo_wari_combo.setItemText(i, en_name if idx == 0 else mr_name)

    def generate_pass_photo_pdf(self):
        idx = 0 if self.lang == "en" else 1
        wari_name = self.pass_photo_wari_combo.currentData()
        wari_year = self.pass_photo_year_spin.value()

        self._set_busy(True, LABELS["status_loading"][idx])
        try:
            wari_sheet = get_sheet(WARI_SHEET_NAME)
            wari_values = wari_sheet.get_all_values()
        except Exception as e:
            self._set_busy(False)
            QMessageBox.critical(self, LABELS["error_title"][idx], f"Could not read Wari Attendees: {e}")
            return

        attendee_card_ids = [
            row[2] for row in wari_values[1:]
            if len(row) >= 3 and row[0] == wari_name and str(row[1]) == str(wari_year) and row[2].strip()
        ]
        if not attendee_card_ids:
            self._set_busy(False)
            QMessageBox.warning(self, LABELS["pdf_no_attendees_title"][idx], LABELS["pdf_no_attendees_body"][idx])
            return

        try:
            details_map = self.fetch_full_details_map()
        except Exception as e:
            self._set_busy(False)
            QMessageBox.critical(self, LABELS["error_title"][idx], f"Could not read Sevak Details: {e}")
            return
        self._set_busy(False)

        default_name = f"Wari_{wari_name}_{wari_year}_PassPhotos.pdf"
        save_path, _ = QFileDialog.getSaveFileName(
            self, LABELS["pass_pdf_save_dialog_title"][idx], default_name, "PDF Files (*.pdf)"
        )
        if not save_path:
            return

        self._set_busy(True, LABELS["status_generating_pdf"][idx])
        try:
            self._render_pass_photo_pdf(save_path, attendee_card_ids, details_map)
        except Exception as e:
            self._set_busy(False)
            QMessageBox.critical(self, LABELS["error_title"][idx], f"Failed to create PDF: {e}")
            return
        self._set_busy(False)
        QMessageBox.information(self, LABELS["success_title"][idx], LABELS["pdf_success_body"][idx])

    def _fit_pass_photo_name(self, last_name, first_name, font_name, font_size, max_width):
        """Returns "Last First" if it fits within max_width, otherwise
        abbreviates the first name to its initial + a dot ("Last F.")."""
        last_name = (last_name or "").strip()
        first_name = (first_name or "").strip()
        full = " ".join(filter(None, [last_name, first_name]))
        if pdfmetrics.stringWidth(full, font_name, font_size) <= max_width or not first_name:
            return full
        abbreviated = " ".join(filter(None, [last_name, f"{first_name[0]}."]))
        return abbreviated

    def _render_pass_photo_pdf(self, save_path, card_ids, details_map):
        """Renders a grid of passport-style photos (4 per row, 5 rows per
        page = 20/page) with the member's name below each, for printing on
        glossy paper and cutting out to stick on an ID card."""
        PHOTO_W = 2.8 * cm
        PHOTO_H = 3 * cm
        NAME_H = 0.6 * cm
        CELL_W = PHOTO_W
        CELL_H = PHOTO_H + NAME_H
        COLS = 4
        ROWS_PER_PAGE = 5
        PER_PAGE = COLS * ROWS_PER_PAGE
        NAME_FONT_SIZE = 8
        # Gap between rows so each row has clear space above and below it
        # once cut out, and a small gap between columns so side-by-side
        # cuts don't need to split a shared border either.
        ROW_GAP = 0.3 * cm
        COL_GAP = 0.1 * cm

        page_width, page_height = A4
        grid_w = COLS * CELL_W + (COLS - 1) * COL_GAP
        grid_h = ROWS_PER_PAGE * CELL_H + (ROWS_PER_PAGE - 1) * ROW_GAP
        margin_left = (page_width - grid_w) / 2
        margin_top = (page_height - grid_h) / 2

        c = pdfcanvas.Canvas(save_path, pagesize=A4)
        photo_cache = {}

        for i, card_id in enumerate(card_ids):
            record = details_map.get(card_id, {})
            pos_in_page = i % PER_PAGE
            if i > 0 and pos_in_page == 0:
                c.showPage()

            row = pos_in_page // COLS
            col = pos_in_page % COLS
            x_left = margin_left + col * (CELL_W + COL_GAP)
            y_top = page_height - margin_top - row * (CELL_H + ROW_GAP)
            y_bottom = y_top - CELL_H
            photo_y_bottom = y_top - PHOTO_H

            # border around the whole cell (photo + name)
            c.rect(x_left, y_bottom, CELL_W, CELL_H)
            # divider between photo and name area
            c.line(x_left, photo_y_bottom, x_left + CELL_W, photo_y_bottom)

            photo_path = self._resolve_photo_for_pdf(record.get("photo", ""), photo_cache)
            if photo_path:
                try:
                    c.drawImage(
                        photo_path, x_left, photo_y_bottom, width=PHOTO_W, height=PHOTO_H,
                        preserveAspectRatio=False, mask="auto",
                    )
                except Exception:
                    pass

            last_en = record.get("last_en", "") or record.get("last_mr", "")
            first_en = record.get("first_en", "") or record.get("first_mr", "")
            name_text = self._fit_pass_photo_name(
                last_en, first_en, "Helvetica", NAME_FONT_SIZE, CELL_W - 0.1 * cm
            )
            c.setFont("Helvetica", NAME_FONT_SIZE)
            c.drawCentredString(
                x_left + CELL_W / 2, y_bottom + NAME_H / 2 - NAME_FONT_SIZE / 3, name_text
            )

        c.save()

    # ---------- Import Records page ----------

    # Target fields offered for column mapping, in display order, reusing
    # the same label keys as the Add Member form.
    IMPORT_FIELDS = [
        ("card_id", "lbl_card_id"), ("first_en", "lbl_first_en"), ("middle_en", "lbl_middle_en"),
        ("last_en", "lbl_last_en"), ("dob", "lbl_dob"), ("phone", "lbl_phone"),
        ("address_en", "lbl_address_en"), ("mukkam_en", "lbl_mukkam_en"), ("post_en", "lbl_post_en"),
        ("taluka_en", "lbl_taluka_en"), ("district_en", "lbl_district_en"), ("state_en", "lbl_state_en"),
        ("pincode", "lbl_pincode"),
        ("first_mr", "lbl_first_mr"), ("middle_mr", "lbl_middle_mr"), ("last_mr", "lbl_last_mr"),
        ("address_mr", "lbl_address_mr"), ("mukkam_mr", "lbl_mukkam_mr"), ("post_mr", "lbl_post_mr"),
        ("taluka_mr", "lbl_taluka_mr"), ("jilha_mr", "lbl_jilha_mr"), ("state_mr", "lbl_state_mr"),
    ]
    # Fields whose values should run through the legacy-font converter when
    # the "legacy font" checkbox is on (Marathi text columns only).
    IMPORT_MARATHI_FIELDS = {
        "first_mr", "middle_mr", "last_mr", "address_mr",
        "mukkam_mr", "post_mr", "taluka_mr", "jilha_mr", "state_mr",
    }

    def _build_import_records_page(self):
        page = QWidget()
        outer_layout = QVBoxLayout()
        page.setLayout(outer_layout)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_content = QWidget()
        scroll.setWidget(scroll_content)
        layout = QVBoxLayout()
        scroll_content.setLayout(layout)
        outer_layout.addWidget(scroll)

        file_row = QHBoxLayout()
        self.import_select_file_btn = QPushButton()
        self.import_select_file_btn.clicked.connect(self.select_import_excel_file)
        self.import_file_label = QLabel("(no file selected)")
        file_row.addWidget(self.import_select_file_btn)
        file_row.addWidget(self.import_file_label, 1)
        layout.addLayout(file_row)

        sheet_row = QHBoxLayout()
        self.import_sheet_label = QLabel()
        self.import_sheet_combo = QComboBox()
        self.import_sheet_combo.currentIndexChanged.connect(self._on_import_sheet_changed)
        sheet_row.addWidget(self.import_sheet_label)
        sheet_row.addWidget(self.import_sheet_combo, 1)
        layout.addLayout(sheet_row)

        self.import_legacy_font_checkbox = QCheckBox()
        layout.addWidget(self.import_legacy_font_checkbox)

        self.import_mapping_label = QLabel()
        layout.addWidget(self.import_mapping_label)

        mapping_form = QFormLayout()
        self.import_field_labels = {}
        self.import_column_combos = {}
        for field_key, label_key in self.IMPORT_FIELDS:
            label = QLabel()
            self.import_field_labels[field_key] = label
            combo = QComboBox()
            self.import_column_combos[field_key] = combo
            mapping_form.addRow(label, combo)
        layout.addLayout(mapping_form)

        btn_row = QHBoxLayout()
        self.import_preview_btn = QPushButton()
        self.import_preview_btn.clicked.connect(self.preview_import_records)
        self.import_all_btn = QPushButton()
        self.import_all_btn.setStyleSheet("font-size: 14px; padding: 10px; background-color: #4CAF50; color: white;")
        self.import_all_btn.clicked.connect(self.import_all_records)
        btn_row.addWidget(self.import_preview_btn)
        btn_row.addWidget(self.import_all_btn)
        layout.addLayout(btn_row)

        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        layout.addWidget(separator)

        self.bulk_translate_btn = QPushButton()
        self.bulk_translate_btn.setStyleSheet(
            "font-size: 14px; padding: 10px; background-color: #16a085; color: white;"
        )
        self.bulk_translate_btn.clicked.connect(self.start_bulk_translate)
        layout.addWidget(self.bulk_translate_btn)

        self.import_preview_table = QTableWidget()
        layout.addWidget(self.import_preview_table, 1)

        return page

    def _refresh_import_field_labels(self):
        idx = 0 if self.lang == "en" else 1
        for field_key, label_key in self.IMPORT_FIELDS:
            self.import_field_labels[field_key].setText(LABELS[label_key][idx])

    def select_import_excel_file(self):
        filepath, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File", "", "Excel Files (*.xlsx *.xlsm *.xls)"
        )
        if not filepath:
            return
        idx = 0 if self.lang == "en" else 1
        self._set_busy(True, LABELS["status_loading"][idx])
        try:
            workbook = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        except Exception as e:
            self._set_busy(False)
            QMessageBox.critical(self, LABELS["error_title"][idx], f"Could not open file: {e}")
            return
        self._set_busy(False)

        self.import_workbook_path = filepath
        self.import_file_label.setText(os.path.basename(filepath))
        self.import_sheet_combo.blockSignals(True)
        self.import_sheet_combo.clear()
        self.import_sheet_combo.addItems(workbook.sheetnames)
        self.import_sheet_combo.blockSignals(False)
        self._on_import_sheet_changed()

    def _on_import_sheet_changed(self):
        sheet_name = self.import_sheet_combo.currentText()
        if not sheet_name or not getattr(self, "import_workbook_path", None):
            return
        idx = 0 if self.lang == "en" else 1
        self._set_busy(True, LABELS["status_loading"][idx])
        try:
            workbook = openpyxl.load_workbook(self.import_workbook_path, data_only=True, read_only=True)
            ws = workbook[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            self._set_busy(False)
            QMessageBox.critical(self, LABELS["error_title"][idx], f"Could not read sheet: {e}")
            return
        self._set_busy(False)

        self.import_header = [str(c).strip() if c is not None else "" for c in (rows[0] if rows else [])]
        self.import_data_rows = rows[1:] if len(rows) > 1 else []
        self._populate_import_column_combos()

    def _populate_import_column_combos(self):
        idx = 0 if self.lang == "en" else 1
        none_label = LABELS["(none)"][idx]
        options = [none_label] + self.import_header
        for field_key, label_key in self.IMPORT_FIELDS:
            combo = self.import_column_combos[field_key]
            combo.blockSignals(True)
            combo.clear()
            combo.addItems(options)
            best_match = self._auto_match_column(field_key, label_key)
            if best_match is not None:
                combo.setCurrentIndex(self.import_header.index(best_match) + 1)
            else:
                combo.setCurrentIndex(0)
            combo.blockSignals(False)

    def _auto_match_column(self, field_key, label_key):
        """Best-effort auto-match of a target field to an Excel header, by
        loose case-insensitive matching against the field key and its
        bilingual label text."""
        candidates = {field_key.lower(), field_key.replace("_", " ").lower()}
        for lang_idx in (0, 1):
            label_text = LABELS[label_key][lang_idx]
            candidates.add(label_text.lower().replace(":", "").strip())
        for header in self.import_header:
            h = header.lower().strip()
            if h in candidates:
                return header
        for header in self.import_header:
            h = header.lower().strip()
            if any(c and (c in h or h in c) for c in candidates):
                return header
        return None

    def _mapped_row_dict(self, row):
        """Builds a {field_key: value} dict for one Excel data row, applying
        the legacy-font converter to Marathi fields if that option is on."""
        convert = self.import_legacy_font_checkbox.isChecked()
        result = {}
        for field_key, _ in self.IMPORT_FIELDS:
            combo = self.import_column_combos[field_key]
            col_name = combo.currentText()
            if not col_name or col_name in (LABELS["(none)"][0], LABELS["(none)"][1]):
                result[field_key] = ""
                continue
            col_index = self.import_header.index(col_name)
            value = row[col_index] if col_index < len(row) else ""
            value = "" if value is None else str(value).strip()
            if convert and field_key in self.IMPORT_MARATHI_FIELDS and value:
                value = convert_legacy_marathi(value)
            result[field_key] = value
        return result

    def preview_import_records(self):
        idx = 0 if self.lang == "en" else 1
        if not getattr(self, "import_data_rows", None):
            QMessageBox.warning(self, LABELS["import_no_rows_title"][idx], LABELS["import_no_rows_body"][idx])
            return

        preview_rows = self.import_data_rows[:10]
        mapped = [self._mapped_row_dict(row) for row in preview_rows]
        field_keys = [fk for fk, _ in self.IMPORT_FIELDS]

        self.import_preview_table.clear()
        self.import_preview_table.setColumnCount(len(field_keys))
        self.import_preview_table.setHorizontalHeaderLabels(
            [LABELS[lk][idx] for _, lk in self.IMPORT_FIELDS]
        )
        self.import_preview_table.setRowCount(len(mapped))
        for r, record in enumerate(mapped):
            for c, key in enumerate(field_keys):
                self.import_preview_table.setItem(r, c, QTableWidgetItem(record.get(key, "")))

    def import_all_records(self):
        idx = 0 if self.lang == "en" else 1
        if not getattr(self, "import_data_rows", None):
            QMessageBox.warning(self, LABELS["import_no_rows_title"][idx], LABELS["import_no_rows_body"][idx])
            return

        card_id_combo = self.import_column_combos["card_id"]
        if card_id_combo.currentIndex() == 0:
            QMessageBox.warning(self, LABELS["import_no_card_id_title"][idx], LABELS["import_no_card_id_body"][idx])
            return

        confirm = QMessageBox.question(
            self, LABELS["import_confirm_title"][idx],
            f"{LABELS['import_confirm_title'][idx]}: {len(self.import_data_rows)} "
            f"{'rows' if idx == 0 else 'नोंदी'}?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
        )
        if confirm != QMessageBox.Yes:
            return

        self._set_busy(True, LABELS["status_saving"][idx])
        imported = 0
        try:
            sheet = get_sheet()
            header_map = get_header_map(sheet)
            width = (max(header_map.values()) + 1) if header_map else len(DETAILS_FIELDS)
            existing_rows = len(sheet.get_all_values())
            new_rows = []
            for i, row in enumerate(self.import_data_rows):
                record = self._mapped_row_dict(row)
                if not record.get("card_id"):
                    continue
                record["serial_no"] = existing_rows + len(new_rows)
                for numeric_key in ("card_id", "phone", "pincode"):
                    if numeric_key in record:
                        record[numeric_key] = _to_number(record[numeric_key])
                out_row = [""] * width
                for key, value in record.items():
                    col_i = header_map.get(DETAILS_FIELDS.get(key, ""))
                    if col_i is not None:
                        out_row[col_i] = value
                new_rows.append(out_row)
            if new_rows:
                sheet.append_rows(new_rows)
                imported = len(new_rows)
        except Exception as e:
            self._set_busy(False)
            QMessageBox.critical(self, LABELS["error_title"][idx], f"Import failed: {e}")
            return
        self._set_busy(False)
        QMessageBox.information(
            self, LABELS["success_title"][idx],
            f"{imported} {'records imported successfully!' if idx == 0 else 'नोंदी यशस्वीरित्या आयात झाल्या!'}"
        )

    def start_bulk_translate(self):
        idx = 0 if self.lang == "en" else 1
        confirm = QMessageBox.question(
            self, LABELS["bulk_translate_confirm_title"][idx], LABELS["bulk_translate_confirm_body"][idx],
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
        )
        if confirm != QMessageBox.Yes:
            return

        self._set_busy(True, LABELS["status_translating"][idx])
        self._bulk_translate_worker = BulkTranslateWorker()
        self._bulk_translate_worker.progress.connect(self._on_bulk_translate_progress)
        self._bulk_translate_worker.finished_ok.connect(self._on_bulk_translate_finished)
        self._bulk_translate_worker.failed.connect(self._on_bulk_translate_failed)
        self._bulk_translate_worker.start()

    def _on_bulk_translate_progress(self, current, total):
        idx = 0 if self.lang == "en" else 1
        self.statusBar().showMessage(f"{LABELS['status_translating'][idx]} ({current}/{total})")

    def _on_bulk_translate_finished(self, translated_count, photo_linked_count):
        idx = 0 if self.lang == "en" else 1
        self._set_busy(False)
        body = (
            f"{translated_count} Marathi field(s) translated and {photo_linked_count} photo link(s) added!"
            if idx == 0
            else f"{translated_count} मराठी फील्ड भाषांतरित आणि {photo_linked_count} फोटो लिंक जोडल्या!"
        )
        QMessageBox.information(self, LABELS["success_title"][idx], body)

    def _on_bulk_translate_failed(self, message):
        idx = 0 if self.lang == "en" else 1
        self._set_busy(False)
        QMessageBox.critical(self, LABELS["error_title"][idx], message)

    def fetch_full_details_map(self):
        """Reads the Sevak Details sheet into a dict keyed by card_id, with
        every field needed to render the Wari photo list."""
        sheet = get_sheet(DETAILS_SHEET_NAME)
        values = sheet.get_all_values()
        if not values:
            return {}
        header_map = get_header_map(sheet)
        details = {}
        for row in values[1:]:
            record = row_to_details_record(row, header_map)
            if not record["card_id"]:
                continue
            details[record["card_id"]] = record
        return details

    # Supersampling factor for shaped-text rasterization: glyphs are
    # rendered at font_size_pt * SHAPED_TEXT_SUPERSAMPLE pixel resolution,
    # then scaled back down to the correct physical point size when placed
    # in the PDF. Without this the raster is as coarse as the tiny point
    # size itself and looks blurry/pixelated once drawn.
    SHAPED_TEXT_SUPERSAMPLE = 4

    @staticmethod
    def _draw_shaped_line(c, text, x, baseline_y, font_size_pt):
        """Draws one line of (possibly Marathi) text using proper OpenType
        shaping, so conjuncts/matras render correctly instead of as
        separate unjoined marks. Falls back to plain drawString if shaping
        fails for any reason (e.g. pure ASCII, or font/shaping unavailable)."""
        text = (text or "").strip()
        if not text:
            return
        scale = SevakJodaForm.SHAPED_TEXT_SUPERSAMPLE
        try:
            result = render_text_line(text, _PDF_FONT_PATH, font_size_pt * scale)
        except Exception:
            result = None

        if result is None:
            c.setFont(PDF_FONT_NAME, font_size_pt)
            c.drawString(x, baseline_y, text)
            return

        mask_img, ascent_px = result
        # mask_img is a coverage mask (0 = no ink, 255 = full ink); composite
        # it as black text onto a white background for the PDF.
        from PIL import Image as _PILImage
        composed = _PILImage.new("RGB", mask_img.size, (255, 255, 255))
        composed.paste((0, 0, 0), (0, 0), mask_img)

        # scale raster pixel dimensions back down to true point size
        width_pt = composed.width / scale
        height_pt = composed.height / scale
        ascent_pt = ascent_px / scale
        image_bottom_y = baseline_y - (height_pt - ascent_pt)
        c.drawImage(
            ImageReader(composed), x, image_bottom_y,
            width=width_pt, height=height_pt, mask=None,
        )

    def _resolve_photo_for_pdf(self, photo_value, cache):
        """Returns a local file path for the given photo value (Drive link
        or local path), downloading from Drive if needed. Caches by value
        so the same photo isn't re-downloaded for every attendee."""
        if not photo_value:
            return None
        if photo_value in cache:
            return cache[photo_value]

        local_path = None
        drive_id = self._extract_drive_file_id(photo_value)
        if drive_id:
            try:
                dest = os.path.join(PHOTOS_DIR, f"pdf_cache_{drive_id}.jpg")
                if not os.path.exists(dest):
                    drive_helper.download_photo(drive_id, dest)
                local_path = dest
            except RuntimeError:
                local_path = None
        elif os.path.exists(photo_value):
            local_path = photo_value

        cache[photo_value] = local_path
        return local_path

    @staticmethod
    def _build_mukkam_post_line(mukkam_mr, post_mr):
        """मु. <mukkam> पो. <post>, or मु. पो. <post> if mukkam is blank.
        Returns "" if both are blank (so the caller can skip the line)."""
        mukkam_mr = (mukkam_mr or "").strip()
        post_mr = (post_mr or "").strip()
        if not mukkam_mr and not post_mr:
            return ""
        if not mukkam_mr:
            return f"मु. पो. {post_mr}".strip()
        text = f"मु. {mukkam_mr}"
        if post_mr:
            text += f" पो. {post_mr}"
        return text

    @staticmethod
    def _build_taluka_jilha_line(taluka_mr, jilha_mr):
        """ता. <taluka> जि. <jilha>. Returns "" if both are blank."""
        parts = []
        if (taluka_mr or "").strip():
            parts.append(f"ता. {taluka_mr.strip()}")
        if (jilha_mr or "").strip():
            parts.append(f"जि. {jilha_mr.strip()}")
        return " ".join(parts)

    @staticmethod
    def _build_state_pincode_line(state_mr, pincode):
        """<state> <pincode>, no prefix on either. Returns "" if both blank."""
        parts = [p.strip() for p in (state_mr, pincode) if (p or "").strip()]
        return " ".join(parts)

    def _build_photo_list_block_lines(self, full_name, record, phone):
        """Builds the block's text lines: name, then up to 4 address-derived
        lines (address / mukkam+post / taluka+jilha / state+pincode - each
        skipped entirely if blank, never left as an empty line), then phone."""
        lines = [full_name]

        address_mr = (record.get("address_mr", "") or "").strip()
        if address_mr:
            lines.append(address_mr)

        mukkam_post = self._build_mukkam_post_line(record.get("mukkam_mr", ""), record.get("post_mr", ""))
        if mukkam_post:
            lines.append(mukkam_post)

        taluka_jilha = self._build_taluka_jilha_line(record.get("taluka_mr", ""), record.get("jilha_mr", ""))
        if taluka_jilha:
            lines.append(taluka_jilha)

        state_pincode = self._build_state_pincode_line(record.get("state_mr", ""), record.get("pincode", ""))
        if state_pincode:
            lines.append(state_pincode)

        phone = (phone or "").strip()
        if phone:
            lines.append(f"मोबाईल: {phone}")
        return lines

    def generate_wari_photo_list_pdf(self):
        idx = 0 if self.lang == "en" else 1
        wari_name = self.photo_list_wari_combo.currentData()
        wari_year = self.photo_list_year_spin.value()

        self._set_busy(True, LABELS["status_loading"][idx])
        try:
            wari_sheet = get_sheet(WARI_SHEET_NAME)
            wari_values = wari_sheet.get_all_values()
        except Exception as e:
            self._set_busy(False)
            QMessageBox.critical(self, LABELS["error_title"][idx], f"Could not read Wari Attendees: {e}")
            return

        attendee_card_ids = [
            row[2] for row in wari_values[1:]
            if len(row) >= 3 and row[0] == wari_name and str(row[1]) == str(wari_year) and row[2].strip()
        ]
        if not attendee_card_ids:
            self._set_busy(False)
            QMessageBox.warning(self, LABELS["pdf_no_attendees_title"][idx], LABELS["pdf_no_attendees_body"][idx])
            return

        try:
            details_map = self.fetch_full_details_map()
        except Exception as e:
            self._set_busy(False)
            QMessageBox.critical(self, LABELS["error_title"][idx], f"Could not read Sevak Details: {e}")
            return
        self._set_busy(False)

        default_name = f"Wari_{wari_name}_{wari_year}_PhotoList.pdf"
        save_path, _ = QFileDialog.getSaveFileName(
            self, LABELS["pdf_save_dialog_title"][idx], default_name, "PDF Files (*.pdf)"
        )
        if not save_path:
            return

        self._set_busy(True, LABELS["status_generating_pdf"][idx])
        try:
            self._render_photo_list_pdf(save_path, attendee_card_ids, details_map)
        except Exception as e:
            self._set_busy(False)
            QMessageBox.critical(self, LABELS["error_title"][idx], f"Failed to create PDF: {e}")
            return
        self._set_busy(False)
        QMessageBox.information(self, LABELS["success_title"][idx], LABELS["pdf_success_body"][idx])

    # Column widths for the Wari photo list PDF. Sr.no / Name / Photo per
    # the original spec (1.5 + 6.7 + 2.8 = 11cm); Pass no. / Sign are new,
    # blank columns appended for handwriting.
    PHOTO_LIST_COL1_W = 1.5 * cm   # Sr. no.
    PHOTO_LIST_COL2_W = 6.7 * cm   # Name (+ address block)
    PHOTO_LIST_COL3_W = 2.8 * cm   # Photo
    PHOTO_LIST_COL4_W = 1.5 * cm   # Pass no.
    PHOTO_LIST_COL5_W = 1.5 * cm   # Sign
    PHOTO_LIST_HEADERS = ["Sr. no.", "Name", "Photo", "Pass no.", "Sign"]

    def _draw_photo_list_page_header(self, c, x_left, y_top, col_widths, header_h, font_size):
        """Draws the repeating column-header row at the top of each page."""
        total_w = sum(col_widths)
        y_bottom = y_top - header_h
        c.rect(x_left, y_bottom, total_w, header_h)
        c.setFont(PDF_FONT_NAME, font_size)

        x = x_left
        for width, label in zip(col_widths, self.PHOTO_LIST_HEADERS):
            if x > x_left:
                c.line(x, y_bottom, x, y_top)
            c.drawCentredString(x + width / 2, y_bottom + header_h / 2 - font_size / 3, label)
            x += width

    def _render_photo_list_pdf(self, save_path, card_ids, details_map):
        COL1_W = self.PHOTO_LIST_COL1_W
        COL2_W = self.PHOTO_LIST_COL2_W
        COL3_W = self.PHOTO_LIST_COL3_W
        COL4_W = self.PHOTO_LIST_COL4_W
        COL5_W = self.PHOTO_LIST_COL5_W
        col_widths = [COL1_W, COL2_W, COL3_W, COL4_W, COL5_W]
        BLOCK_W = sum(col_widths)
        BLOCK_H = 3 * cm
        HEADER_H = 0.4 * cm
        MEMBERS_PER_PAGE = 7
        MARGIN_TOP = 1.5 * cm
        MARGIN_LEFT = 2 * cm

        page_width, page_height = A4
        c = pdfcanvas.Canvas(save_path, pagesize=A4)
        photo_cache = {}
        FONT_SIZE = 9
        HEADER_FONT_SIZE = 8

        def new_page_header():
            self._draw_photo_list_page_header(
                c, MARGIN_LEFT, page_height - MARGIN_TOP, col_widths, HEADER_H, HEADER_FONT_SIZE
            )

        new_page_header()

        for i, card_id in enumerate(card_ids):
            record = details_map.get(card_id, {})
            pos_in_page = i % MEMBERS_PER_PAGE
            if i > 0 and pos_in_page == 0:
                c.showPage()
                new_page_header()

            y_top = page_height - MARGIN_TOP - HEADER_H - pos_in_page * BLOCK_H
            y_bottom = y_top - BLOCK_H
            x_left = MARGIN_LEFT

            # block + column borders
            c.rect(x_left, y_bottom, BLOCK_W, BLOCK_H)
            x = x_left
            for width in col_widths[:-1]:
                x += width
                c.line(x, y_bottom, x, y_top)

            c.setFont(PDF_FONT_NAME, FONT_SIZE)

            # column 1: serial number, near the top of the cell with a small gap
            SERIAL_TOP_GAP = 0.2 * cm
            c.drawCentredString(
                x_left + COL1_W / 2, y_top - SERIAL_TOP_GAP - FONT_SIZE, str(i + 1)
            )

            # column 2: Marathi name / address / mukkam+post / taluka+jilha / state+pincode / phone
            # Falls back to the English name (then, only as a last resort,
            # the raw card_id) if Marathi translation was never filled in.
            full_name = (
                " ".join(filter(None, [
                    record.get("last_mr", ""), record.get("first_mr", ""), record.get("middle_mr", ""),
                ]))
                or " ".join(filter(None, [
                    record.get("last_en", ""), record.get("first_en", ""), record.get("middle_en", ""),
                ]))
                or record.get("card_id", card_id)
            )
            lines = self._build_photo_list_block_lines(full_name, record, record.get("phone", ""))
            text_x = x_left + COL1_W + 0.15 * cm
            line_height = BLOCK_H / max(6.0, len(lines) + 0.2)
            text_y = y_top - line_height * 0.8
            for line in lines:
                self._draw_shaped_line(c, line, text_x, text_y, FONT_SIZE)
                text_y -= line_height

            # column 3: photo, stretched to fill the cell edge-to-edge
            photo_path = self._resolve_photo_for_pdf(record.get("photo", ""), photo_cache)
            if photo_path:
                try:
                    col3_x = x_left + COL1_W + COL2_W
                    c.drawImage(
                        photo_path, col3_x, y_bottom, width=COL3_W, height=BLOCK_H,
                        preserveAspectRatio=False, mask="auto",
                    )
                except Exception:
                    pass

            # columns 4 (Pass no.) and 5 (Sign) are left blank for handwriting

        c.save()

    def _mr_row(self, field):
        """Wraps a Marathi input with a small button that opens the
        on-screen Marathi keyboard targeting that field."""
        row = QWidget()
        row_layout = QHBoxLayout()
        row_layout.setContentsMargins(0, 0, 0, 0)
        kb_btn = QPushButton("⌨")
        kb_btn.setFixedWidth(30)
        kb_btn.clicked.connect(lambda: self.open_marathi_keyboard(field))
        row_layout.addWidget(field)
        row_layout.addWidget(kb_btn, 0, Qt.AlignTop)
        row.setLayout(row_layout)
        return row

    def open_marathi_keyboard(self, field):
        self.marathi_keyboard.attach_to(field)
        self.marathi_keyboard.show()
        self.marathi_keyboard.raise_()
        self.marathi_keyboard.activateWindow()

    def translate_fields(self, pairs=None):
        """Translates every English name/address field into Marathi,
        always overwriting whatever is currently in the Marathi field.
        Defaults to the Add Member page's fields if no pairs are given.
        Runs the network calls on a background thread so a slow/blocked
        connection can't freeze the UI."""
        pairs = pairs if pairs is not None else self.translation_pairs
        texts_by_index = {}
        for i, (en_field, mr_field) in enumerate(pairs):
            en_text = _get_text(en_field).strip()
            if en_text:
                texts_by_index[i] = en_text
        if not texts_by_index:
            return

        idx = 0 if self.lang == "en" else 1
        self._active_translate_pairs = pairs
        self._set_busy(True, LABELS["status_translating"][idx])

        self._translate_worker = TranslateWorker(texts_by_index)
        self._translate_worker.finished_ok.connect(self._on_translate_finished)
        self._translate_worker.failed.connect(self._on_translate_failed)
        self._translate_worker.start()

    def _on_translate_finished(self, results):
        self._set_busy(False)
        pairs = self._active_translate_pairs
        for i, translated in results.items():
            _, mr_field = pairs[i]
            _set_text(mr_field, translated)

    def _on_translate_failed(self, message):
        self._set_busy(False)
        QMessageBox.warning(self, "Translation Error", message)

    def _current_card_id(self, context):
        field = self.card_id if context == "add" else self.edit_card_id
        return field.text().strip()

    def _rename_photo_to_card_id(self, filepath, context):
        """Renames a just-captured photo file to <card_id>.<ext>, if a
        card_id has already been entered. Leaves it untouched otherwise."""
        card_id = self._current_card_id(context)
        if not card_id:
            return filepath
        safe_card_id = "".join(c for c in card_id if c.isalnum() or c in ("-", "_")).strip()
        if not safe_card_id:
            return filepath
        ext = os.path.splitext(filepath)[1].lower() or ".jpg"
        new_path = os.path.join(PHOTOS_DIR, f"{safe_card_id}{ext}")
        try:
            if os.path.abspath(new_path) != os.path.abspath(filepath):
                if os.path.exists(new_path):
                    os.remove(new_path)
                shutil.move(filepath, new_path)
            return new_path
        except OSError:
            return filepath

    def run_scan(self, context="add"):
        """Triggers the WIA scanner dialog and captures the photo."""
        idx = 0 if self.lang == "en" else 1
        self._set_busy(True, LABELS["status_scanning"][idx])
        try:
            filepath = scan_photo()
            filepath = self._rename_photo_to_card_id(filepath, context)
        except RuntimeError as e:
            self._set_busy(False)
            QMessageBox.warning(self, "Scanner Error", str(e))
            return
        except Exception as e:
            self._set_busy(False)
            QMessageBox.critical(self, "Scanner Error", f"Scan failed: {e}")
            return
        self.set_photo(filepath, context=context, upload_to_drive=True)
        self._set_busy(False)

    def browse_local_photo(self, context="add"):
        """Lets the user pick an existing image file from disk instead of scanning."""
        filepath, _ = QFileDialog.getOpenFileName(
            self, "Select Photo", "", "Images (*.png *.jpg *.jpeg *.bmp)"
        )
        if not filepath:
            return
        idx = 0 if self.lang == "en" else 1
        self._set_busy(True, LABELS["status_uploading"][idx])
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ext = os.path.splitext(filepath)[1].lower() or ".jpg"
            dest_path = os.path.join(PHOTOS_DIR, f"photo_{timestamp}{ext}")
            shutil.copy(filepath, dest_path)
        except Exception as e:
            self._set_busy(False)
            QMessageBox.critical(self, "File Error", f"Could not load photo: {e}")
            return
        self.set_photo(dest_path, context=context, upload_to_drive=True)
        self._set_busy(False)

    def browse_drive_photo(self, context="add"):
        """Lets the user pick an existing photo already stored in the Drive folder."""
        idx = 0 if self.lang == "en" else 1
        self._set_busy(True, LABELS["status_loading"][idx])
        try:
            files = drive_helper.list_photos()
        except RuntimeError as e:
            self._set_busy(False)
            QMessageBox.critical(self, LABELS["error_title"][idx], str(e))
            return
        self._set_busy(False)

        dialog = QDialog(self)
        dialog.setWindowTitle(LABELS["drive_picker_title"][idx])
        dialog.resize(360, 420)
        layout = QVBoxLayout()
        dialog.setLayout(layout)

        listbox = QListWidget()
        if not files:
            item = QListWidgetItem(LABELS["drive_no_photos"][idx])
            item.setFlags(Qt.NoItemFlags)
            listbox.addItem(item)
        else:
            for f in files:
                item = QListWidgetItem(f["name"])
                item.setData(Qt.UserRole, f)
                listbox.addItem(item)
        layout.addWidget(listbox)

        def on_choose(item):
            data = item.data(Qt.UserRole)
            if not data:
                return
            dialog.accept()
            self._load_photo_from_drive(data["id"], data["name"], context=context)

        listbox.itemDoubleClicked.connect(on_choose)
        dialog.exec_()

    def _load_photo_from_drive(self, file_id, name, context="add"):
        idx = 0 if self.lang == "en" else 1
        self._set_busy(True, LABELS["status_loading"][idx])
        try:
            ext = os.path.splitext(name)[1].lower() or ".jpg"
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            dest_path = os.path.join(PHOTOS_DIR, f"drive_{timestamp}{ext}")
            drive_helper.download_photo(file_id, dest_path)
            self.set_photo(dest_path, context=context, upload_to_drive=False)
            prefix = "" if context == "add" else "edit_"
            setattr(self, f"{prefix}photo_drive_id", file_id)
            setattr(self, f"{prefix}photo_drive_link", f"https://drive.google.com/file/d/{file_id}/view")
        except RuntimeError as e:
            self._set_busy(False)
            QMessageBox.critical(self, LABELS["error_title"][idx], str(e))
            return
        self._set_busy(False)

    def set_photo(self, filepath, context="add", upload_to_drive=False):
        idx = 0 if self.lang == "en" else 1
        prefix = "" if context == "add" else "edit_"
        preview = self.photo_preview if context == "add" else self.edit_photo_preview

        setattr(self, f"{prefix}photo_path", filepath)
        setattr(self, f"{prefix}photo_drive_id", None)
        setattr(self, f"{prefix}photo_drive_link", None)
        pixmap = QPixmap(filepath).scaled(
            preview.width(), preview.height(),
            Qt.KeepAspectRatio
        )
        preview.setPixmap(pixmap)

        if upload_to_drive:
            try:
                file_id, link = drive_helper.upload_photo(filepath)
                setattr(self, f"{prefix}photo_drive_id", file_id)
                setattr(self, f"{prefix}photo_drive_link", link)
            except RuntimeError as e:
                self._set_busy(False)
                QMessageBox.warning(
                    self, LABELS["error_title"][idx], f'{LABELS["drive_upload_failed"][idx]}: {e}'
                )

    def save_member(self):
        idx = 0 if self.lang == "en" else 1
        if not self.first_name_en.text().strip():
            QMessageBox.warning(self, LABELS["missing_info_title"][idx], LABELS["missing_info_body"][idx])
            return

        self._set_busy(True, LABELS["status_saving"][idx])
        try:
            card_id = self.card_id.text().strip()

            sheet = get_sheet()
            existing_rows = len(sheet.get_all_values())  # includes header
            serial_no = existing_rows  # header is row 1, so this gives next serial
            header_map = get_header_map(sheet)

            field_values = {
                "serial_no": serial_no,
                "card_id": _to_number(card_id),
                "first_en": self.first_name_en.text().strip(),
                "middle_en": self.middle_name_en.text().strip(),
                "last_en": self.last_name_en.text().strip(),
                "dob": self.dob.text().strip(),
                "phone": _to_number(self.phone_number.text().strip()),
                "address_en": _get_text(self.address_en).strip(),
                "mukkam_en": self.mukkam_en.text().strip(),
                "post_en": self.post_en.text().strip(),
                "taluka_en": self.taluka_en.text().strip(),
                "district_en": self.district_en.text().strip(),
                "state_en": self.state_en.text().strip(),
                "pincode": _to_number(self.pincode.text().strip()),
                "first_mr": self.first_name_mr.text().strip(),
                "middle_mr": self.middle_name_mr.text().strip(),
                "last_mr": self.last_name_mr.text().strip(),
                "address_mr": _get_text(self.address_mr).strip(),
                "mukkam_mr": self.mukkam_mr.text().strip(),
                "post_mr": self.post_mr.text().strip(),
                "taluka_mr": self.taluka_mr.text().strip(),
                "jilha_mr": self.jilha_mr.text().strip(),
                "state_mr": self.state_mr.text().strip(),
                "photo": self.photo_drive_link or self.photo_path or "",
            }

            # build a row sized to the sheet's real width, placing each value
            # at its actual header-matched column instead of a fixed position
            width = (max(header_map.values()) + 1) if header_map else len(field_values)
            row = [""] * width
            for key, value in field_values.items():
                col_i = header_map.get(DETAILS_FIELDS[key])
                if col_i is not None:
                    row[col_i] = value
            sheet.append_row(row)
        except Exception as e:
            self._set_busy(False)
            QMessageBox.critical(self, LABELS["error_title"][idx], f"Failed to save: {e}")
            return
        self._set_busy(False)
        QMessageBox.information(self, LABELS["success_title"][idx], LABELS["success_body"][idx])
        self.clear_form()

    def clear_form(self):
        for field in [self.card_id, self.first_name_en, self.middle_name_en, self.last_name_en, self.dob,
                      self.phone_number, self.address_en, self.mukkam_en, self.post_en, self.taluka_en,
                      self.district_en, self.state_en, self.pincode, self.first_name_mr,
                      self.middle_name_mr, self.last_name_mr, self.address_mr, self.mukkam_mr,
                      self.post_mr, self.taluka_mr, self.jilha_mr, self.state_mr]:
            _clear(field)
        self.photo_path = None
        self.photo_drive_id = None
        self.photo_drive_link = None
        self.photo_preview.clear()
        self.photo_preview.setText(LABELS["no_photo"][0 if self.lang == "en" else 1])

    # ---------- Edit Sevak logic ----------

    @staticmethod
    def _extract_drive_file_id(url):
        match = re.search(r"/d/([a-zA-Z0-9_-]+)", url or "")
        return match.group(1) if match else None

    def load_edit_member(self):
        idx = 0 if self.lang == "en" else 1
        card_id = self.edit_card_id_search.text().strip()
        if not card_id:
            return

        self._set_busy(True, LABELS["status_loading"][idx])
        try:
            sheet = get_sheet(DETAILS_SHEET_NAME)
            values = sheet.get_all_values()
        except Exception as e:
            self._set_busy(False)
            QMessageBox.critical(self, LABELS["error_title"][idx], f"Could not load member: {e}")
            return

        if not values:
            self._set_busy(False)
            QMessageBox.warning(self, LABELS["no_member_found_title"][idx], LABELS["no_member_found_body"][idx])
            return
        header_map = get_header_map(sheet)
        card_id_col = header_map.get(DETAILS_FIELDS["card_id"])

        for row_number, row in enumerate(values[1:], start=2):
            if card_id_col is not None and card_id_col < len(row) and row[card_id_col].strip() == card_id:
                self._populate_edit_fields(row_number, row, header_map)
                self._set_busy(False)
                return

        self._set_busy(False)
        QMessageBox.warning(self, LABELS["no_member_found_title"][idx], LABELS["no_member_found_body"][idx])

    def _populate_edit_fields(self, row_number, row, header_map):
        idx = 0 if self.lang == "en" else 1
        record = row_to_details_record(row, header_map)

        self.edit_row_number = row_number
        self.edit_card_id.setText(record["card_id"])
        self.edit_first_name_en.setText(record["first_en"])
        self.edit_middle_name_en.setText(record["middle_en"])
        self.edit_last_name_en.setText(record["last_en"])
        self.edit_dob.setText(record["dob"])
        self.edit_phone_number.setText(record["phone"])
        _set_text(self.edit_address_en, record["address_en"])
        self.edit_mukkam_en.setText(record["mukkam_en"])
        self.edit_post_en.setText(record["post_en"])
        self.edit_taluka_en.setText(record["taluka_en"])
        self.edit_district_en.setText(record["district_en"])
        self.edit_state_en.setText(record["state_en"])
        self.edit_pincode.setText(record["pincode"])
        self.edit_first_name_mr.setText(record["first_mr"])
        self.edit_middle_name_mr.setText(record["middle_mr"])
        self.edit_last_name_mr.setText(record["last_mr"])
        _set_text(self.edit_address_mr, record["address_mr"])
        self.edit_mukkam_mr.setText(record["mukkam_mr"])
        self.edit_post_mr.setText(record["post_mr"])
        self.edit_taluka_mr.setText(record["taluka_mr"])
        self.edit_jilha_mr.setText(record["jilha_mr"])
        self.edit_state_mr.setText(record["state_mr"])

        photo_value = record["photo"]
        self.edit_original_photo_value = photo_value
        self.edit_photo_path = None
        self.edit_photo_drive_id = None
        self.edit_photo_drive_link = None
        self.edit_photo_preview.clear()

        if photo_value:
            drive_id = self._extract_drive_file_id(photo_value)
            if drive_id:
                try:
                    dest = os.path.join(PHOTOS_DIR, f"edit_preview_{drive_id}.jpg")
                    drive_helper.download_photo(drive_id, dest)
                    self.set_photo(dest, context="edit", upload_to_drive=False)
                    self.edit_photo_drive_id = drive_id
                    self.edit_photo_drive_link = photo_value
                except RuntimeError:
                    self.edit_photo_preview.setText(LABELS["no_photo"][idx])
            elif os.path.exists(photo_value):
                self.set_photo(photo_value, context="edit", upload_to_drive=False)
            else:
                self.edit_photo_preview.setText(LABELS["no_photo"][idx])
        else:
            self.edit_photo_preview.setText(LABELS["no_photo"][idx])

    def update_member(self):
        idx = 0 if self.lang == "en" else 1
        if not self.edit_row_number:
            QMessageBox.warning(self, LABELS["no_member_loaded_title"][idx], LABELS["no_member_loaded_body"][idx])
            return
        if not self.edit_first_name_en.text().strip():
            QMessageBox.warning(self, LABELS["missing_info_title"][idx], LABELS["missing_info_body"][idx])
            return

        self._set_busy(True, LABELS["status_saving"][idx])
        try:
            photo_value = (
                self.edit_photo_drive_link or self.edit_photo_path or self.edit_original_photo_value or ""
            )
            field_values = {
                "card_id": _to_number(self.edit_card_id.text().strip()),
                "first_en": self.edit_first_name_en.text().strip(),
                "middle_en": self.edit_middle_name_en.text().strip(),
                "last_en": self.edit_last_name_en.text().strip(),
                "dob": self.edit_dob.text().strip(),
                "phone": _to_number(self.edit_phone_number.text().strip()),
                "address_en": _get_text(self.edit_address_en).strip(),
                "mukkam_en": self.edit_mukkam_en.text().strip(),
                "post_en": self.edit_post_en.text().strip(),
                "taluka_en": self.edit_taluka_en.text().strip(),
                "district_en": self.edit_district_en.text().strip(),
                "state_en": self.edit_state_en.text().strip(),
                "pincode": _to_number(self.edit_pincode.text().strip()),
                "first_mr": self.edit_first_name_mr.text().strip(),
                "middle_mr": self.edit_middle_name_mr.text().strip(),
                "last_mr": self.edit_last_name_mr.text().strip(),
                "address_mr": _get_text(self.edit_address_mr).strip(),
                "mukkam_mr": self.edit_mukkam_mr.text().strip(),
                "post_mr": self.edit_post_mr.text().strip(),
                "taluka_mr": self.edit_taluka_mr.text().strip(),
                "jilha_mr": self.edit_jilha_mr.text().strip(),
                "state_mr": self.edit_state_mr.text().strip(),
                "photo": photo_value,
            }

            sheet = get_sheet()
            header_map = get_header_map(sheet)

            # update only the specific cells whose headers we recognize,
            # rather than overwriting a fixed column range (which would
            # clobber any columns - like mukkampost/taluka/district - that
            # sit between the fields this app knows about)
            batch_data = []
            for key, value in field_values.items():
                col = details_col_index(header_map, key)
                if col is not None:
                    batch_data.append({
                        "range": gspread.utils.rowcol_to_a1(self.edit_row_number, col),
                        "values": [[value]],
                    })
            if batch_data:
                sheet.batch_update(batch_data)
        except Exception as e:
            self._set_busy(False)
            QMessageBox.critical(self, LABELS["error_title"][idx], f"Failed to update: {e}")
            return
        self._set_busy(False)
        QMessageBox.information(self, LABELS["success_title"][idx], LABELS["update_success_body"][idx])


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = SevakJodaForm()
    window.show()
    sys.exit(app.exec_())
